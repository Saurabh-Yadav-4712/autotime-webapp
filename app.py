from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import string
import random
import os
import smtplib
from email.mime.text import MIMEText

def send_otp_email(to_email, otp, context="Account Verification"):
    smtp_email = os.environ.get('SMTP_EMAIL')
    smtp_password = os.environ.get('SMTP_PASSWORD')
    
    if not smtp_email or not smtp_password:
        print("\n" + "="*50)
        print(f"🔔 [MOCK EMAIL] EMAIL SENT TO: {to_email}")
        print(f"🔔 CONTEXT: {context}")
        print(f"🔔 OTP CODE: {otp}")
        print("="*50 + "\n")
        return False
        
    try:
        msg = MIMEText(f"Your AutoTime {context} OTP is: {otp}")
        msg['Subject'] = f'AutoTime - {context} OTP'
        msg['From'] = smtp_email
        msg['To'] = to_email
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(smtp_email, smtp_password)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False
app = Flask(__name__)
app.config['SECRET_KEY'] = 'super_secret_key_for_autotime'
db_url = os.environ.get('DATABASE_URL', 'sqlite:///autotime.db')
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==========================================
# DATABASE MODELS
# ==========================================
class Institute(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    institute_code = db.Column(db.String(20), unique=True, nullable=False)
    admin_username = db.Column(db.String(50), unique=True, nullable=False)
    admin_email = db.Column(db.String(100), unique=True, nullable=False)
    admin_password = db.Column(db.String(255), nullable=False)

class Teacher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    teacher_id = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    departments = db.Column(db.String(100), nullable=False)
    available_days = db.Column(db.String(100), nullable=False)
    max_hours = db.Column(db.Integer, nullable=False)
    password = db.Column(db.String(255), nullable=True)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    class_id = db.Column(db.String(50), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    division = db.Column(db.String(10), nullable=False)

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    subject_code = db.Column(db.String(20), nullable=False)
    subject_name = db.Column(db.String(100), nullable=False)
    class_id = db.Column(db.String(100), nullable=False)
    teacher_id = db.Column(db.String(20), nullable=False)
    total_course_hours = db.Column(db.Integer, nullable=False, default=50)
    required_hours = db.Column(db.Integer, nullable=False)
    subject_type = db.Column(db.String(50), default='Theory')
    session_length = db.Column(db.Integer, default=1)
    preferred_days = db.Column(db.String(100), default="")

class Timetable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    class_id = db.Column(db.String(50), nullable=False)
    day_name = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.String(20), nullable=False)
    end_time = db.Column(db.String(20), nullable=False)
    subject_name = db.Column(db.String(100), nullable=False)
    teacher_name = db.Column(db.String(100), nullable=False)
    is_proxy = db.Column(db.Boolean, default=False)
class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    key = db.Column(db.String(50), nullable=False)
    value = db.Column(db.String(100), nullable=False)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    institute_code = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    class_id = db.Column(db.String(50), nullable=False)


def generate_institute_code(prefix="INS"):
    chars = string.ascii_uppercase + string.digits
    return f"{prefix}-{''.join(random.choices(chars, k=5))}"
def get_dynamic_time_slots(inst_code):
    # Get settings from DB
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    
    # Fallback default values (agar admin ne settings save nahi ki)
    total_lectures = int(s.get('total_lectures', 4))
    start_time_str = s.get('start_time', '08:00')
    lec_duration = int(s.get('lecture_duration', 45))
    break_duration = int(s.get('break_time', 30))
    lunch_after = int(s.get('lunch_after_lecture', 2))

    time_slots = []
    # Convert string to datetime object for calculation
    current_time = datetime.strptime(start_time_str, '%H:%M')

    for i in range(1, total_lectures + 1):
        end_time = current_time + timedelta(minutes=lec_duration)
        
        # Format time to "08:00 AM" style
        start_str = current_time.strftime('%I:%M %p')
        end_str = end_time.strftime('%I:%M %p')
        
        time_slots.append((start_str, end_str))
        
        current_time = end_time
        
        # Add Lunch Break duration after the specified lecture
        if i == lunch_after:
            current_time = current_time + timedelta(minutes=break_duration)
            
    return time_slots

def trim_time_slots(schedule, time_slots, lunch_after):
    max_slot_index = -1
    for day, slots in schedule.items():
        for i, slot in enumerate(time_slots):
            if slot[0] in slots:
                max_slot_index = max(max_slot_index, i)
    
    if max_slot_index != -1:
        # Keep at least up to lunch_after so break UI doesn't crash
        cutoff = max(max_slot_index + 1, lunch_after)
        return time_slots[:cutoff]
    return time_slots
with app.app_context():
    db.create_all()

# ==========================================
# AUTH & DASHBOARD ROUTES
# ==========================================
@app.route('/')
def home():
    return render_template('main_site/landing.html')

@app.route('/login')
def login_page():
    return render_template('auth/auth.html')

@app.route('/register_institute', methods=['GET', 'POST'])
def register_institute():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].strip()
        college_name = request.form['college_name'].strip()
        password = request.form['password']

        if Institute.query.filter_by(admin_username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('register_institute'))
        if Institute.query.filter_by(admin_email=email).first():
            flash('Email already registered.', 'danger')
            return redirect(url_for('register_institute'))
        
        # Save temp data and send OTP
        otp = str(random.randint(100000, 999999))
        session['reg_data'] = {
            'type': 'institute',
            'college_name': college_name,
            'username': username,
            'email': email,
            'password': password
        }
        session['reg_otp'] = otp
        
        email_sent = send_otp_email(email, otp, context="Institute Registration")
        if not email_sent:
            flash('Error sending email or SMTP not configured. OTP printed in server logs.', 'warning')
        else:
            flash('An OTP has been sent to your email for verification.', 'info')
        return redirect(url_for('verify_reg_otp'))
    return render_template('auth/register_institute.html')

@app.route('/verify_reg_otp', methods=['GET', 'POST'])
def verify_reg_otp():
    if 'reg_data' not in session or 'reg_otp' not in session:
        flash('Session expired. Please register again.', 'danger')
        return redirect(url_for('login_page'))
        
    if request.method == 'POST':
        user_otp = request.form['otp'].strip()
        if user_otp == session['reg_otp']:
            data = session['reg_data']
            
            if data['type'] == 'institute':
                inst_code = generate_institute_code()
                new_institute = Institute(
                    name=data['college_name'],
                    institute_code=inst_code,
                    admin_username=data['username'],
                    admin_email=data['email'],
                    admin_password=generate_password_hash(data['password'])
                )
                db.session.add(new_institute)
                db.session.commit()
                
                # Clear session
                session.pop('reg_data', None)
                session.pop('reg_otp', None)
                
                flash(f'College Registered Successfully! Your Institute Code is: {inst_code}', 'success')
                return redirect(url_for('login_page'))
                
            elif data['type'] == 'student':
                new_student = Student(
                    institute_code=data['inst_code'],
                    name=data['name'],
                    email=data['email'],
                    class_id=data['class_id'],
                    password=generate_password_hash(data['password'])
                )
                db.session.add(new_student)
                db.session.commit()
                
                session.pop('reg_data', None)
                session.pop('reg_otp', None)
                
                flash('Student Registered Successfully! You can now login.', 'success')
                return redirect(url_for('login_page'))
                
        else:
            flash('Invalid OTP. Please try again.', 'danger')
            return redirect(url_for('verify_reg_otp'))
            
    return render_template('auth/verify_reg_otp.html')

@app.route('/login_admin', methods=['GET', 'POST'])
def login_admin():
    if request.method == 'GET':
        return redirect(url_for('login_page'))
        
    admin = Institute.query.filter_by(admin_username=request.form['username']).first()
    if admin and check_password_hash(admin.admin_password, request.form['password']):
        session['admin_id'] = admin.id
        session['institute_code'] = admin.institute_code
        flash(f'Welcome back, {admin.name}!', 'success')
        return redirect(url_for('admin_dash'))
    flash('Invalid Credentials!', 'danger')
    return redirect(url_for('login_page'))

@app.route('/admin_dash')
def admin_dash():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    
    # Calculate Analytics
    c_count = Course.query.filter_by(institute_code=inst_code).count()
    t_count = Teacher.query.filter_by(institute_code=inst_code).count()
    s_count = Subject.query.filter_by(institute_code=inst_code).count()
    
    # Check if timetable generated
    generated = Timetable.query.filter_by(institute_code=inst_code).first() is not None
    
    # Free teachers (rough logic: those whose assigned hours < max_hours)
    # We can approximate free teachers by checking how many teachers have fewer lectures in Timetable than max_hours, or just list the count of teachers.
    # To keep it efficient:
    all_teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    assigned_counts = {t.name: 0 for t in all_teachers}
    
    if generated:
        # Fix: Count unique slots (day + time) to avoid double counting Common subjects
        tt_entries = Timetable.query.filter_by(institute_code=inst_code).all()
        teacher_slots = {}
        for entry in tt_entries:
            base_name = entry.teacher_name.replace(" (Proxy)", "")
            teacher_slots.setdefault(base_name, set()).add((entry.day_name, entry.start_time))
            
        for name, slots in teacher_slots.items():
            if name in assigned_counts:
                assigned_counts[name] += len(slots)
    
    free_teachers_count = 0
    faculty_workload = []
    
    for t in all_teachers:
        assigned = assigned_counts[t.name]
        free_hrs = max(0, t.max_hours - assigned)
        
        if free_hrs > 0:
            free_teachers_count += 1
            
        faculty_workload.append({
            'name': t.name,
            'max_hours': t.max_hours,
            'assigned_hours': assigned,
            'free_hours': free_hrs
        })

    import math
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    weeks_setting = Settings.query.filter_by(institute_code=inst_code, key='weeks_per_semester').first()
    weeks_per_semester = int(weeks_setting.value) if weeks_setting else 15
    
    syllabus_tracking = []
    for s in subjects:
        syllabus_tracking.append({
            'name': s.subject_name,
            'code': s.subject_code,
            'teacher': s.teacher_id,
            'total_hrs': s.total_course_hours,
            'weekly_hrs': s.required_hours,
            'weeks_needed': math.ceil(s.total_course_hours / s.required_hours) if s.required_hours > 0 else 0
        })

    admin = Institute.query.get(session['admin_id'])
    return render_template('admin/admin_dash.html', admin=admin, c_count=c_count, t_count=t_count, s_count=s_count, generated=generated, free_teachers=free_teachers_count, faculty_workload=faculty_workload, syllabus_tracking=syllabus_tracking, weeks_per_semester=weeks_per_semester)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ==========================================
# MANAGEMENT ROUTES
# ==========================================
@app.route('/manage_courses', methods=['GET', 'POST'])
def manage_courses():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        db.session.add(Course(
            institute_code=inst_code, class_id=request.form['class_id'],
            department=request.form['department'], semester=request.form['semester'],
            division=request.form['division']
        ))
        db.session.commit()
        flash('Course added!', 'success')
        return redirect(url_for('manage_courses'))
    
    courses = Course.query.filter_by(institute_code=inst_code).all()
    return render_template('admin/manage_master.html', manage_type='course', items=courses)

@app.route('/manage_teachers', methods=['GET', 'POST'])
def manage_teachers():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        if Teacher.query.filter_by(email=request.form['email']).first():
            flash('Email already registered!', 'danger')
            return redirect(url_for('manage_teachers'))
            
        days = request.form.getlist('days')
        if not days:
            flash('Select at least one day!', 'danger')
            return redirect(url_for('manage_teachers'))

        db.session.add(Teacher(
            institute_code=inst_code, teacher_id=request.form['teacher_id'],
            name=request.form['name'], email=request.form['email'],
            departments=request.form['departments'], available_days=",".join(days),
            max_hours=int(request.form['max_hours'])
        ))
        db.session.commit()
        flash('Teacher added!', 'success')
        return redirect(url_for('manage_teachers'))
    
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    unique_depts = list(set([c.department for c in courses]))
    return render_template('admin/manage_master.html', manage_type='teacher', items=teachers, depts=unique_depts)

@app.route('/manage_subjects', methods=['GET', 'POST'])
def manage_subjects():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        class_ids = request.form.getlist('class_id')
        if not class_ids:
            flash('Select at least one Class!', 'danger')
            return redirect(url_for('manage_subjects'))

        weeks_setting = Settings.query.filter_by(institute_code=inst_code, key='weeks_per_semester').first()
        weeks = int(weeks_setting.value) if weeks_setting else 15
        total_hours = int(request.form['total_course_hours'])
        import math
        
        db.session.add(Subject(
            institute_code=inst_code, subject_code=request.form['subject_code'],
            subject_name=request.form['subject_name'], class_id=",".join(class_ids),
            teacher_id=request.form['teacher_id'], total_course_hours=total_hours,
            required_hours=math.ceil(total_hours / weeks),
            subject_type=request.form['subject_type'], session_length=int(request.form.get('session_length', 1))
        ))
        db.session.commit()
        flash('Subject mapped!', 'success')
        return redirect(url_for('manage_subjects'))
    
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    return render_template('admin/manage_master.html', manage_type='subject', items=subjects, courses=courses, teachers=teachers)

# ==========================================
# EDIT ROUTES
# ==========================================
@app.route('/edit_course/<int:id>', methods=['GET', 'POST'])
def edit_course(id):
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    course = Course.query.get_or_404(id)
    if request.method == 'POST':
        course.class_id = request.form['class_id']
        course.department = request.form['department']
        course.semester = request.form['semester']
        course.division = request.form['division']
        db.session.commit()
        flash('Course updated!', 'success')
        return redirect(url_for('manage_courses'))
    return render_template('admin/edit_master.html', item=course, edit_type='course')

@app.route('/edit_teacher/<int:id>', methods=['GET', 'POST'])
def edit_teacher(id):
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    teacher = Teacher.query.get_or_404(id)
    if request.method == 'POST':
        teacher.teacher_id = request.form['teacher_id']
        teacher.name = request.form['name']
        teacher.email = request.form['email']
        teacher.departments = request.form['departments']
        teacher.max_hours = request.form['max_hours']
        days = request.form.getlist('days')
        if days: teacher.available_days = ",".join(days)
        db.session.commit()
        flash('Teacher updated!', 'success')
        return redirect(url_for('manage_teachers'))
    
    courses = Course.query.filter_by(institute_code=session['institute_code']).all()
    depts = list(set([c.department for c in courses]))
    return render_template('admin/edit_master.html', item=teacher, edit_type='teacher', unique_depts=depts)

@app.route('/edit_subject/<int:id>', methods=['GET', 'POST'])
def edit_subject(id):
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    subject = Subject.query.get_or_404(id)
    if request.method == 'POST':
        subject.subject_code = request.form['subject_code']
        subject.subject_name = request.form['subject_name']
        class_ids = request.form.getlist('class_id')
        if class_ids: subject.class_id = ",".join(class_ids)
        subject.teacher_id = request.form['teacher_id']
        weeks_setting = Settings.query.filter_by(institute_code=session['institute_code'], key='weeks_per_semester').first()
        weeks = int(weeks_setting.value) if weeks_setting else 15
        total_hours = int(request.form['total_course_hours'])
        import math
        subject.total_course_hours = total_hours
        subject.required_hours = math.ceil(total_hours / weeks)
        subject.subject_type = request.form['subject_type']
        subject.session_length = int(request.form.get('session_length', 1))
        db.session.commit()
        flash('Subject updated!', 'success')
        return redirect(url_for('manage_subjects'))
    
    courses = Course.query.filter_by(institute_code=session['institute_code']).all()
    teachers = Teacher.query.filter_by(institute_code=session['institute_code']).all()
    return render_template('admin/edit_master.html', item=subject, edit_type='subject', courses=courses, teachers=teachers)

# ==========================================
# DELETE ROUTES
# ==========================================
@app.route('/delete/<type>/<int:id>')
def delete_item(type, id):
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    
    if type == 'course':
        item = Course.query.get_or_404(id)
        route = 'manage_courses'
    elif type == 'teacher':
        item = Teacher.query.get_or_404(id)
        route = 'manage_teachers'
    else:
        item = Subject.query.get_or_404(id)
        route = 'manage_subjects'

    if item.institute_code == session['institute_code']:
        db.session.delete(item)
        db.session.commit()
        flash('Deleted successfully.', 'info')
    
    return redirect(url_for(route))
# ==========================================
# TIMETABLE ENGINE & ALGORITHM
# ==========================================


# ==========================================
# ⚡ THE MASTER ALGORITHM (TIMETABLE ENGINE)
# ==========================================
@app.route('/generate_timetable', methods=['GET', 'POST'])
def generate_timetable():
    if request.method == 'GET':
        return redirect(url_for('admin_dash'))
        
    if 'admin_id' not in session:
        return redirect(url_for('login_page'))
        
    inst_code = session['institute_code']
    
    # 🧹 1. Clear existing timetable for this institute (Fresh Start)
    Timetable.query.filter_by(institute_code=inst_code).delete()
    db.session.commit()
    
    # 📥 2. Fetch all required Data
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    
    # Dictionary for quick Teacher lookups
    teacher_dict = {t.teacher_id: t for t in teachers}
    
    # Get Time Slots (Assuming your get_dynamic_time_slots function is present)
    time_slots = get_dynamic_time_slots(inst_code) 
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    
    # 🧠 3. Initialize Tracking States (To prevent Clashes)
    # Track which class is busy at what time
    class_timetable = {c.class_id: {day: {} for day in days} for c in courses}
    # Track which teacher is busy at what time
    teacher_timetable = {t.teacher_id: {day: {} for day in days} for t in teachers}
    # Track maximum hours given to a teacher
    teacher_hours = {t.teacher_id: 0 for t in teachers}
    
    # ⚔️ Phase Separation
    common_subjects = []
    normal_subjects = []
    
    for sub in subjects:
        if ',' in sub.class_id:
            common_subjects.append(sub)  # e.g. "FYCS, FYIT, FYBMS"
        else:
            normal_subjects.append(sub)

    # 🚀 PHASE 1: Process Common Subjects (Highest Priority)
    import random
    random.shuffle(common_subjects)
    for sub in common_subjects:
        # Split and clean class IDs
        target_classes = [c.strip() for c in sub.class_id.split(',')]
        assigned_hours = 0
        teacher = teacher_dict.get(sub.teacher_id)
        
        while assigned_hours < sub.required_hours:
            scheduled_this_round = False
            # Randomness + Balanced Load
            random.shuffle(days)
            days.sort(key=lambda d: max(len(class_timetable.get(c, {}).get(d, {})) for c in target_classes))
            
            for day in days:
                if assigned_hours >= sub.required_hours: break
                
                # Constraint: Preferred Days for Subject
                if sub.preferred_days and day not in sub.preferred_days: continue
                
                # Constraint: Is teacher available on this day?
                if teacher and day not in teacher.available_days: continue
                
                # Check for contiguous blocks based on session_length
                valid_indices = list(range(len(time_slots) - sub.session_length + 1))
                
                for idx in valid_indices:
                    slots_to_check = time_slots[idx : idx + sub.session_length]
                    
                    classes_free = all(s[0] not in class_timetable.get(c, {}).get(day, {}) for c in target_classes for s in slots_to_check)
                    teacher_free = all(s[0] not in teacher_timetable.get(sub.teacher_id, {}).get(day, {}) for s in slots_to_check)
                    hours_ok = teacher_hours.get(sub.teacher_id, 0) + sub.session_length <= teacher.max_hours if teacher else True
                    
                    if classes_free and teacher_free and hours_ok:
                        # ✅ ASSIGN TO ALL CLASSES SIMULTANEOUSLY
                        for s in slots_to_check:
                            for c in target_classes:
                                if c in class_timetable:
                                    class_timetable[c][day][s[0]] = (sub, s[1])
                            teacher_timetable[sub.teacher_id][day][s[0]] = (sub, s[1])
                            
                        if teacher: teacher_hours[sub.teacher_id] += sub.session_length
                        assigned_hours += sub.session_length
                        scheduled_this_round = True
                        break # Move to next day (Horizontal distribution)
                        
            if not scheduled_this_round: break # Stuck

    # 🚀 PHASE 2: Process Normal Subjects
    random.shuffle(normal_subjects)
    for sub in normal_subjects:
        assigned_hours = 0
        target_class = sub.class_id
        teacher = teacher_dict.get(sub.teacher_id)
        
        if target_class not in class_timetable: continue
        
        while assigned_hours < sub.required_hours:
            scheduled_this_round = False
            # Randomness + Balanced Load
            random.shuffle(days)
            days.sort(key=lambda d: len(class_timetable[target_class][d]))
            
            for day in days:
                if assigned_hours >= sub.required_hours: break
                
                # Constraint: Preferred Days for Subject
                if sub.preferred_days and day not in sub.preferred_days: continue
                
                # Constraint: Is teacher available on this day?
                if teacher and day not in teacher.available_days: continue
                
                valid_indices = list(range(len(time_slots) - sub.session_length + 1))
                
                for idx in valid_indices:
                    slots_to_check = time_slots[idx : idx + sub.session_length]
                    
                    # CLASH DETECTION
                    class_free = all(s[0] not in class_timetable[target_class][day] for s in slots_to_check)
                    teacher_free = all(s[0] not in teacher_timetable.get(sub.teacher_id, {}).get(day, {}) for s in slots_to_check)
                    hours_ok = teacher_hours.get(sub.teacher_id, 0) + sub.session_length <= teacher.max_hours if teacher else True
                    
                    if class_free and teacher_free and hours_ok:
                        # ✅ ASSIGN LECTURE
                        for s in slots_to_check:
                            class_timetable[target_class][day][s[0]] = (sub, s[1])
                            teacher_timetable[sub.teacher_id][day][s[0]] = (sub, s[1])
                            
                        if teacher: teacher_hours[sub.teacher_id] += sub.session_length
                        assigned_hours += sub.session_length
                        scheduled_this_round = True
                        break # Move to next day (Horizontal distribution)
                        
            if not scheduled_this_round: break # Stuck
    # Removed Phase 3 (Aggressive Compaction) because it broke session_length blocks.

    # 🚀 PHASE 4: Cross-Day Gap Elimination (Move isolated lectures to other days)
    for c_id in class_timetable.keys():
        for _ in range(3): # Multiple passes to resolve cascading gaps
            for day in days:
                slots_data = class_timetable[c_id].get(day, {})
                if not slots_data: continue
                
                filled_indices = [i for i, slot in enumerate(time_slots) if slot[0] in slots_data]
                if not filled_indices: continue
                
                if max(filled_indices) >= len(filled_indices):
                    # Gap detected! Find the lectures placed AFTER the gap
                    sorted_filled = sorted(filled_indices)
                    for i, idx in enumerate(sorted_filled):
                        if idx > i: # This lecture is separated by a gap
                            st_time = time_slots[idx][0]
                            sub_data = class_timetable[c_id][day][st_time]
                            sub = sub_data[0]
                            
                            # Do not attempt to move common subjects or block subjects (session_length > 1)
                            if ',' in sub.class_id or sub.session_length > 1: continue
                            
                            moved = False
                            for other_day in days:
                                if other_day == day: continue
                                
                                # Find first empty slot on other_day
                                other_slots_data = class_timetable[c_id].get(other_day, {})
                                target_st, target_end = None, None
                                for slot in time_slots:
                                    if slot[0] not in other_slots_data:
                                        target_st, target_end = slot[0], slot[1]
                                        break
                                        
                                if target_st and target_st not in teacher_timetable.get(sub.teacher_id, {}).get(other_day, {}):
                                    # Move successful!
                                    del class_timetable[c_id][day][st_time]
                                    del teacher_timetable[sub.teacher_id][day][st_time]
                                    
                                    class_timetable[c_id][other_day][target_st] = (sub, target_end)
                                    teacher_timetable.setdefault(sub.teacher_id, {}).setdefault(other_day, {})[target_st] = (sub, target_end)
                                    moved = True
                                    break
                            
                            if moved:
                                break # Restart pass for this class since timetable mutated

    # 💾 4. Save the Final Generated Output to Database
    records_to_add = []
    for c_id, days_data in class_timetable.items():
        for day, slots_data in days_data.items():
            for start_time, sub_data in slots_data.items():
                
                sub = sub_data[0]
                end_time = sub_data[1]
                
                teacher = teacher_dict.get(sub.teacher_id)
                t_name = teacher.name if teacher else sub.teacher_id
                
                # Create Database Entry
                new_entry = Timetable(
                    institute_code=inst_code,
                    class_id=c_id,
                    day_name=day,
                    start_time=start_time,
                    end_time=end_time,
                    subject_name=sub.subject_name,
                    teacher_name=t_name,
                    is_proxy=False  # Proxy is always False during master generation
                )
                records_to_add.append(new_entry)
                
    db.session.bulk_save_objects(records_to_add)
    db.session.commit()
    
    flash('⚡ Timetable Generated Successfully! Zero Clashes Detected.', 'success')
    return redirect(url_for('admin_dash'))
# ==========================================
# VIEW TIMETABLE ROUTE (UPDATED)
# ==========================================
@app.route('/view_timetable')
def view_timetable():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    
    courses = Course.query.filter_by(institute_code=inst_code).all()
    selected_class = request.args.get('class_id')
    
    schedule = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    time_slots = get_dynamic_time_slots(inst_code)
    
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    lunch_after = int(s.get('lunch_after_lecture', 2))
    break_duration = int(s.get('break_time', 30)) # Fetch break duration
    
    if selected_class:
        entries = Timetable.query.filter_by(institute_code=inst_code, class_id=selected_class).all()
        for day in days:
            schedule[day] = {}
            for entry in entries:
                if entry.day_name == day:
                    schedule[day][entry.start_time] = entry
        time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    inst = Institute.query.filter_by(institute_code=inst_code).first()
    inst_name = inst.name if inst else "Institute"
    
    # Fetch teachers and subjects for manual editing
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    if selected_class:
        subjects = Subject.query.filter_by(institute_code=inst_code, class_id=selected_class).all()
    else:
        subjects = []
    
    return render_template('shared/view_timetable.html', courses=courses, selected_class=selected_class, schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, break_duration=break_duration, inst_name=inst_name, teachers=teachers, subjects=subjects)

# ==========================================
# EDIT TIMETABLE SLOT (MANUAL)
# ==========================================
@app.route('/edit_timetable_slot', methods=['POST'])
def edit_timetable_slot():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    
    entry_id = request.form.get('entry_id')
    new_subject = request.form.get('new_subject')
    new_teacher = request.form.get('new_teacher')
    
    entry = Timetable.query.get(entry_id)
    if not entry:
        flash('Timetable slot not found.', 'danger')
        return redirect(url_for('view_timetable', class_id=request.form.get('class_id')))
        
    # Check for teacher clash (is the new teacher busy at this exact day and time?)
    if new_teacher != entry.teacher_name:
        clash = Timetable.query.filter_by(
            institute_code=entry.institute_code,
            day_name=entry.day_name,
            start_time=entry.start_time,
            teacher_name=new_teacher
        ).first()
        
        if clash:
            flash(f'Cannot assign {new_teacher}. They are already teaching Class {clash.class_id} at this time!', 'danger')
            return redirect(url_for('view_timetable', class_id=entry.class_id))
            
    # Update entry
    entry.subject_name = new_subject
    entry.teacher_name = new_teacher
    # Clear proxy flag if any
    if hasattr(entry, 'is_proxy') and entry.is_proxy:
        entry.is_proxy = False
        entry.original_teacher = None
        
    db.session.commit()
    flash('Timetable slot updated successfully!', 'success')
    return redirect(url_for('view_timetable', class_id=entry.class_id))

# ==========================================
# ORIGINAL GRID EXCEL EXPORT (OPENPYXL)
# ==========================================
@app.route('/export_timetables')
def export_timetables():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    
    # Fetch Classes
    classes = db.session.query(Timetable.class_id).filter_by(institute_code=inst_code).distinct().all()
    class_ids = [c[0] for c in classes]
    
    if not class_ids:
        flash('No timetable generated yet to export!', 'warning')
        return redirect(url_for('view_timetable'))

    # Fetch Settings & Slots
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    
    institute_name = s.get('institute_name', 'INSTITUTE TIMETABLE').upper()
    lunch_after = int(s.get('lunch_after_lecture', 2))
    
    # Time slots format: "08:00 AM - 09:00 AM"
    time_slots = get_dynamic_time_slots(inst_code)
    slots = [f"{slot[0]} - {slot[1]}" for slot in time_slots]
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

    # Excel Styling Variables
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    lunch_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cls in class_ids:
        safe_sheet_name = str(cls).replace('/', '-').replace('\\', '-')[:31]
        ws = wb.create_sheet(title=safe_sheet_name)
        
        # Titles
        c1 = ws.cell(row=1, column=1, value=institute_name)
        c1.font = title_font; c1.alignment = center_align
        c2 = ws.cell(row=2, column=1, value=f"TIMETABLE / SCHEDULE - {cls}")
        c2.font = header_font; c2.alignment = center_align; c2.fill = blue_fill
        
        ws.cell(row=3, column=1, value="Day").font = header_font
        ws.cell(row=3, column=1).alignment = center_align; ws.cell(row=3, column=1).border = thin_border
        ws.column_dimensions['A'].width = 12
        
        col_idx = 2
        lunch_printed = False
        slot_to_col = {}
        
        # Build Header & Lunch Break Column
        for i, slot in enumerate(slots):
            # Print Lunch Column exactly after the specified lecture
            if not lunch_printed and i == lunch_after:
                lc = ws.cell(row=3, column=col_idx, value="LUNCH\nBREAK")
                lc.font = header_font; lc.alignment = center_align; lc.border = thin_border; lc.fill = lunch_fill
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10
                
                # Merge and rotate text vertically
                ws.merge_cells(start_row=4, start_column=col_idx, end_row=4+len(days)-1, end_column=col_idx)
                lb_cell = ws.cell(row=4, column=col_idx, value="L U N C H   B R E A K")
                lb_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center', wrap_text=True)
                lb_cell.font = Font(name='Arial', size=14, bold=True, color="595959")
                lb_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                for r in range(4, 4+len(days)): ws.cell(row=r, column=col_idx).border = thin_border
                col_idx += 1
                lunch_printed = True
            
            # Print Time Slot Header
            slot_to_col[slot] = col_idx
            cell = ws.cell(row=3, column=col_idx, value=slot.replace(" - ", "\n"))
            cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
            col_idx += 1
            
        max_col = col_idx - 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            
        # Fetch data for this specific class
        tt_entries = Timetable.query.filter_by(institute_code=inst_code, class_id=cls).all()
        tt_dict = {(e.day_name, f"{e.start_time} - {e.end_time}"): f"{e.subject_name}\n({e.teacher_name})" for e in tt_entries}
        
        # Populate Grid Data
        for row_idx, day in enumerate(days, start=4):
            day_cell = ws.cell(row=row_idx, column=1, value=day)
            day_cell.font = header_font; day_cell.alignment = center_align; day_cell.border = thin_border
            
            for slot in slots:
                c_idx = slot_to_col[slot]
                cell = ws.cell(row=row_idx, column=c_idx, value=tt_dict.get((day, slot), "---"))
                cell.font = cell_font; cell.alignment = center_align; cell.border = thin_border
                
    # Save Excel to memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="AutoTime_Master_Schedule.xlsx", as_attachment=True)
# ==========================================
# COLLEGE SETTINGS ROUTE
# ==========================================
@app.route('/college_settings', methods=['GET', 'POST'])
def college_settings():
    if 'admin_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        # Yahan humne 'lunch_after_lecture' add kar diya hai
        keys = ['total_lectures', 'start_time', 'lecture_duration', 'break_time', 'lunch_after_lecture', 'weeks_per_semester']
        for key in keys:
            val = request.form.get(key)
            if val:
                setting = Settings.query.filter_by(institute_code=inst_code, key=key).first()
                if setting:
                    setting.value = val 
                else:
                    new_setting = Settings(institute_code=inst_code, key=key, value=val)
                    db.session.add(new_setting)
        
        db.session.commit()
        flash('College Settings Updated Successfully!', 'success')
        return redirect(url_for('college_settings'))

    all_settings = Settings.query.filter_by(institute_code=inst_code).all()
    settings_dict = {s.key: s.value for s in all_settings}

    return render_template('admin/college_settings.html', settings=settings_dict)
# ==========================================
# STUDENT PORTAL ROUTE
# ==========================================
@app.route('/student_portal')
def student_portal():
    inst_code = request.args.get('inst_code')
    class_id = request.args.get('class_id')
    
    schedule = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    time_slots = []
    lunch_after = 2
    
    if inst_code and class_id:
        # Check if institute exists
        institute = Institute.query.filter_by(institute_code=inst_code).first()
        if not institute:
            flash('Invalid Institute Code!', 'danger')
            return redirect(url_for('student_portal'))
            
        time_slots = get_dynamic_time_slots(inst_code)
        settings = Settings.query.filter_by(institute_code=inst_code).all()
        s = {st.key: st.value for st in settings}
        lunch_after = int(s.get('lunch_after_lecture', 2))
        
        entries = Timetable.query.filter_by(institute_code=inst_code, class_id=class_id).all()
        for day in days:
            schedule[day] = {}
            for entry in entries:
                if entry.day_name == day:
                    schedule[day][entry.start_time] = entry
                    
        time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    return render_template('student/student_portal.html', schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, inst_code=inst_code, class_id=class_id)

# ==========================================
# TEACHER PORTAL ROUTE
# ==========================================
@app.route('/teacher_portal')
def teacher_portal():
    inst_code = request.args.get('inst_code')
    teacher_id = request.args.get('teacher_id')
    
    schedule = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    time_slots = []
    lunch_after = 2
    teacher_name = ""
    
    if inst_code and teacher_id:
        teacher = Teacher.query.filter_by(institute_code=inst_code, teacher_id=teacher_id).first()
        if not teacher:
            flash('Invalid Institute Code or Teacher ID!', 'danger')
            return redirect(url_for('teacher_portal'))
            
        teacher_name = teacher.name
        time_slots = get_dynamic_time_slots(inst_code)
        settings = Settings.query.filter_by(institute_code=inst_code).all()
        s = {st.key: st.value for st in settings}
        lunch_after = int(s.get('lunch_after_lecture', 2))
        
        # Fetch timetable specifically for this teacher
        entries = Timetable.query.filter_by(institute_code=inst_code, teacher_name=teacher_name).all()
        for day in days:
            schedule[day] = {}
            for entry in entries:
                if entry.day_name == day:
                    # Teacher ko dikhna chahiye ki kis Class me lecture hai
                    schedule[day][entry.start_time] = entry

    return render_template('teacher_portal.html', schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, inst_code=inst_code, teacher_id=teacher_id, teacher_name=teacher_name)
# ==========================================
# TEACHER OTP ACTIVATION ROUTES
# ==========================================
@app.route('/activate_teacher', methods=['GET', 'POST'])
def activate_teacher():
    if request.method == 'POST':
        email = request.form['email'].strip()
        inst_code = request.form['inst_code'].strip()
        
        # Check agar is email aur code se koi teacher registered hai
        teacher = Teacher.query.filter_by(email=email, institute_code=inst_code).first()
        
        if not teacher:
            flash('No teacher found with this Email and Institute Code.', 'danger')
            return redirect(url_for('activate_teacher'))
            
        if teacher.password:
            flash('Account is already activated! Please login.', 'warning')
            return redirect(url_for('login_page'))

        # 6-Digit OTP Generate karo
        otp = str(random.randint(100000, 999999))
        
        # Session me temporary save karo verify karne ke liye
        session['activation_email'] = email
        session['activation_inst_code'] = inst_code
        session['activation_otp'] = otp
        
        # Send Real Email
        email_sent = send_otp_email(email, otp, context="Teacher Activation")
        
        if email_sent:
            flash('An OTP has been sent to your email address.', 'success')
        else:
            flash('Error sending email or SMTP not configured. OTP printed in server logs.', 'warning')
        return redirect(url_for('verify_teacher_otp'))
        
    return render_template('teacher/activate_teacher.html')


@app.route('/verify_teacher_otp', methods=['GET', 'POST'])
def verify_teacher_otp():
    if 'activation_email' not in session:
        return redirect(url_for('activate_teacher'))
        
    if request.method == 'POST':
        user_otp = request.form['otp'].strip()
        new_password = request.form['new_password'].strip()
        
        # Verify OTP
        if user_otp == session.get('activation_otp'):
            email = session.get('activation_email')
            inst_code = session.get('activation_inst_code')
            
            teacher = Teacher.query.filter_by(email=email, institute_code=inst_code).first()
            if teacher:
                # Password hash karke save karo
                teacher.password = generate_password_hash(new_password)
                db.session.commit()
                
                # Session se OTP data delete karo (Security)
                session.pop('activation_otp', None)
                session.pop('activation_email', None)
                session.pop('activation_inst_code', None)
                
                flash('Account Activated Successfully! You can now login.', 'success')
                return redirect(url_for('login_page'))
        else:
            flash('Invalid OTP! Please try again.', 'danger')
            
    return render_template('teacher/verify_teacher_otp.html')

# ==========================================
# TEACHER LOGIN & SECURE DASHBOARD
# ==========================================
@app.route('/login_teacher', methods=['GET', 'POST'])
def login_teacher():
    if request.method == 'GET':
        return redirect(url_for('login_page'))
        
    teacher = Teacher.query.filter_by(email=request.form['email']).first()
    if teacher and teacher.password and check_password_hash(teacher.password, request.form['password']):
        session['teacher_id'] = teacher.teacher_id
        session['institute_code'] = teacher.institute_code
        session['teacher_name'] = teacher.name
        session['teacher_dept'] = teacher.departments
        flash(f'Welcome to your portal, Prof. {teacher.name}!', 'success')
        return redirect(url_for('teacher_dash'))
    
    flash('Invalid Email/Password or Account not activated via OTP!', 'danger')
    return redirect(url_for('login_page'))

@app.route('/teacher_dash')
def teacher_dash():
    if 'teacher_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    t_name = session['teacher_name']
    
    # Get dynamic time slots & settings
    time_slots = get_dynamic_time_slots(inst_code)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    lunch_after = int(s.get('lunch_after_lecture', 2))
    
    # 1. Fetch this specific teacher's schedule
    schedule = {day: {} for day in days}
    entries = Timetable.query.filter(
        Timetable.institute_code == inst_code,
        Timetable.teacher_name.like(f"%{t_name}%") # Matches original and proxy names
    ).all()
    
    for entry in entries:
        schedule[entry.day_name][entry.start_time] = entry

    time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    # 2. Fetch all courses for the Read-Only Viewer
    courses = Course.query.filter_by(institute_code=inst_code).all()
    
    today_str = datetime.now().strftime('%a')
    
    inst = Institute.query.filter_by(institute_code=inst_code).first()
    inst_name = inst.name if inst else "Institute"
    
    return render_template('teacher/teacher_dash.html', schedule=schedule, courses=courses, days=days, time_slots=time_slots, lunch_after=lunch_after, t_name=t_name, today_str=today_str, inst_name=inst_name)

@app.route('/teacher_view_class')
def teacher_view_class():
    if 'teacher_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    class_id = request.args.get('class_id')
    
    if not class_id:
        flash('Please select a class to view.', 'warning')
        return redirect(url_for('teacher_dash'))
        
    schedule = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    time_slots = get_dynamic_time_slots(inst_code)
    
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    lunch_after = int(s.get('lunch_after_lecture', 2))
    
    entries = Timetable.query.filter_by(institute_code=inst_code, class_id=class_id).all()
    for day in days:
        schedule[day] = {}
        for entry in entries:
            if entry.day_name == day:
                schedule[day][entry.start_time] = entry
                
    time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    # We can reuse the student_portal template or view_timetable for a read-only view. 
    # Let's render a custom small view or just pass it to view_timetable context since it handles schedule rendering nicely.
    # Actually, we can just use the student_portal template but modify it slightly or just pass it as is.
    return render_template('student/student_portal.html', schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, inst_code=inst_code, class_id=class_id, teacher_view=True)

# ==========================================
# THE AUTO-PROXY ALGORITHM 🌟
# ==========================================
@app.route('/apply_leave', methods=['GET', 'POST'])
def apply_leave():
    if request.method == 'GET':
        return redirect(url_for('teacher_dash'))
        
    if 'teacher_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    leave_day = request.form['leave_day']
    leave_time = request.form.get('leave_time', 'ALL')
    # Step 0: The 1-Hour Rule Restriction
    today_str = datetime.now().strftime('%a')
    if leave_day == today_str:
        setting = Settings.query.filter_by(institute_code=inst_code, key='start_time').first()
        start_time_str = setting.value if setting else '08:00'
        start_time_obj = datetime.strptime(start_time_str, '%H:%M')
        cutoff_time = (start_time_obj - timedelta(hours=1)).time()
        
        if datetime.now().time() > cutoff_time:
            flash(f'Leave for today must be applied before {cutoff_time.strftime("%I:%M %p")}. Please contact Admin.', 'danger')
            return redirect(url_for('teacher_dash'))

    t_name = session['teacher_name']
    t_dept = session['teacher_dept'].split(',')
    
    # Step 1: Find all lectures of this teacher on the leave day
    query = Timetable.query.filter_by(institute_code=inst_code, teacher_name=t_name, day_name=leave_day)
    if leave_time != 'ALL':
        query = query.filter_by(start_time=leave_time)
        
    my_lectures = query.all()
    
    if not my_lectures:
        time_text = f"at {leave_time}" if leave_time != 'ALL' else "on this day"
        flash(f'You have no lectures scheduled {time_text}!', 'info')
        return redirect(url_for('teacher_dash'))
        
    # Step 2: Calculate Current Workload for Workload Balancing
    from sqlalchemy import func
    all_teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    assigned_counts = {t.name: 0 for t in all_teachers}
    
    # Fix: Count unique slots (day + time) to avoid double counting Common subjects
    tt_entries = Timetable.query.filter_by(institute_code=inst_code).all()
    teacher_slots = {}
    for entry in tt_entries:
        base_name = entry.teacher_name.replace(" (Proxy)", "")
        teacher_slots.setdefault(base_name, set()).add((entry.day_name, entry.start_time))
        
    for name, slots in teacher_slots.items():
        if name in assigned_counts:
            assigned_counts[name] += len(slots)
            
    success_count = 0
    
    for lecture in my_lectures:
        free_candidates = []
        
        for candidate in all_teachers:
            if candidate.name == t_name: continue # Skip self
            if leave_day not in candidate.available_days: continue
            
            # BUG FIX: Max Hours Check (Prevents assigning beyond capacity)
            if assigned_counts.get(candidate.name, 0) >= candidate.max_hours: continue
            
            # Check if candidate is FREE at this exact time slot
            is_busy = Timetable.query.filter_by(institute_code=inst_code, teacher_name=candidate.name, day_name=leave_day, start_time=lecture.start_time).first()
            if not is_busy:
                # Priority factor: Same department gets a virtual reduction in workload to prioritize them slightly
                is_same_dept = any(d.strip() in candidate.departments for d in t_dept)
                virtual_workload = assigned_counts[candidate.name] - (5 if is_same_dept else 0)
                free_candidates.append((virtual_workload, candidate))
                
        if free_candidates:
            # Sort by workload (lowest first)
            free_candidates.sort(key=lambda x: x[0])
            replacement = free_candidates[0][1]
            
            # Step 3: Assign Proxy and Update Database (Dynamic Subject)
            lecture.teacher_name = f"{replacement.name} (Proxy)"
            lecture.is_proxy = True
            
            # Dynamic Subject Replacement
            # Check if proxy teaches this class
            matching_sub = Subject.query.filter_by(institute_code=inst_code, teacher_id=replacement.teacher_id).filter(Subject.class_id.contains(lecture.class_id)).first()
            if matching_sub:
                lecture.subject_name = matching_sub.subject_name
            else:
                lecture.subject_name = "Supervision (Proxy)"
                
            # Increment their workload tracking so subsequent proxies in this loop know they are busier
            assigned_counts[replacement.name] += 1
            success_count += 1
            
    db.session.commit()
    time_text = f"for {leave_time}" if leave_time != 'ALL' else f"for {leave_day}"
    flash(f'Leave applied {time_text}! Auto-Proxy successfully assigned replacements for {success_count} lectures.', 'success')
    return redirect(url_for('teacher_dash'))
# ==========================================
# STUDENT REGISTRATION, LOGIN & DASHBOARD
# ==========================================
@app.route('/api/get_classes/<inst_code>')
def get_classes(inst_code):
    courses = Course.query.filter_by(institute_code=inst_code).all()
    classes = [c.class_id for c in courses]
    return jsonify({'classes': classes})

@app.route('/register_student', methods=['GET', 'POST'])
def register_student():
    if request.method == 'POST':
        inst_code = request.form['inst_code'].strip()
        name = request.form['name'].strip()
        email = request.form['email'].strip()
        password = request.form['password'].strip()
        class_id = request.form.get('class_id', '').strip().upper()

        # Institute check
        institute = Institute.query.filter_by(institute_code=inst_code).first()
        if not institute:
            flash('Invalid Institute Code!', 'danger')
            return redirect(url_for('register_student'))
            
        # Class check
        course = Course.query.filter_by(institute_code=inst_code, class_id=class_id).first()
        if not course:
            flash(f'Invalid Class ID for Institute {inst_code}!', 'danger')
            return redirect(url_for('register_student'))

        # Email check
        if Student.query.filter_by(email=email).first():
            flash('Email already registered! Please login.', 'warning')
            return redirect(url_for('login_page'))

        # Generate OTP and store in session
        otp = str(random.randint(100000, 999999))
        session['reg_data'] = {
            'type': 'student',
            'inst_code': inst_code,
            'name': name,
            'email': email,
            'class_id': class_id,
            'password': password
        }
        session['reg_otp'] = otp
        
        email_sent = send_otp_email(email, otp, context="Student Registration")
        if not email_sent:
            flash('Error sending email or SMTP not configured. OTP printed in server logs.', 'warning')
        else:
            flash('An OTP has been sent to your email for verification.', 'info')
        return redirect(url_for('verify_reg_otp'))

    return render_template('auth/register_student.html')

@app.route('/login_student', methods=['GET', 'POST'])
def login_student():
    if request.method == 'GET':
        return redirect(url_for('login_page'))
        
    student = Student.query.filter_by(email=request.form['email']).first()
    if student and check_password_hash(student.password, request.form['password']):
        session['student_id'] = student.id
        session['institute_code'] = student.institute_code
        session['student_name'] = student.name
        session['student_class'] = student.class_id
        flash(f'Welcome to your Live Timetable, {student.name}!', 'success')
        return redirect(url_for('student_dash'))
    
    flash('Invalid Email or Password!', 'danger')
    return redirect(url_for('login_page'))

@app.route('/student_dash')
def student_dash():
    if 'student_id' not in session: return redirect(url_for('login_page'))
    inst_code = session['institute_code']
    class_id = session['student_class']
    s_name = session['student_name']
    
    time_slots = get_dynamic_time_slots(inst_code)
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    lunch_after = int(s.get('lunch_after_lecture', 2))
    
    schedule = {day: {} for day in days}
    entries = Timetable.query.filter_by(institute_code=inst_code, class_id=class_id).all()
    
    for entry in entries:
        schedule[entry.day_name][entry.start_time] = entry

    inst = Institute.query.filter_by(institute_code=inst_code).first()
    inst_name = inst.name if inst else "Institute"
    return render_template('student/student_dash.html', schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, class_id=class_id, s_name=s_name, inst_name=inst_name)

# ==========================================
# FORGOT PASSWORD ROUTES
# ==========================================
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        role = request.form.get('role')
        email = request.form.get('email').strip()
        
        user_found = False
        if role == 'admin':
            user = Institute.query.filter_by(admin_email=email).first()
            if user: user_found = True
        elif role == 'teacher':
            user = Teacher.query.filter_by(email=email).first()
            if user: user_found = True
        elif role == 'student':
            user = Student.query.filter_by(email=email).first()
            if user: user_found = True
            
        if not user_found:
            flash(f'No {role} found with this email address.', 'danger')
            return redirect(url_for('forgot_password'))
            
        import random
        otp = str(random.randint(100000, 999999))
        session['reset_email'] = email
        session['reset_role'] = role
        session['reset_otp'] = otp
        
        email_sent = send_otp_email(email, otp, context="Password Reset")
        if email_sent:
            flash('A reset OTP has been sent to your email.', 'success')
        else:
            flash('Error sending email or SMTP not configured. OTP printed in server logs.', 'warning')
            
        return render_template('auth/reset_password.html', email=email)
        
    return render_template('auth/forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'GET':
        return redirect(url_for('login_page'))
        
    if 'reset_email' not in session or 'reset_role' not in session:
        return redirect(url_for('forgot_password'))
        
    otp_input = request.form.get('otp').strip()
    new_password = request.form.get('new_password').strip()
    
    if otp_input != session.get('reset_otp'):
        flash('Invalid OTP!', 'danger')
        return render_template('auth/reset_password.html', email=session.get('reset_email'))
        
    role = session.get('reset_role')
    email = session.get('reset_email')
    
    if role == 'admin':
        user = Institute.query.filter_by(admin_email=email).first()
        if user: user.admin_password = generate_password_hash(new_password)
    elif role == 'teacher':
        user = Teacher.query.filter_by(email=email).first()
        if user: user.password = generate_password_hash(new_password)
    elif role == 'student':
        user = Student.query.filter_by(email=email).first()
        if user: user.password = generate_password_hash(new_password)
        
    db.session.commit()
    session.pop('reset_email', None)
    session.pop('reset_role', None)
    session.pop('reset_otp', None)
    
    flash('Password successfully reset! Please login.', 'success')
    return redirect(url_for('login_page'))



@app.route('/settings')
def settings():
    if 'admin_id' not in session and 'teacher_id' not in session and 'student_id' not in session:
        flash('Please login to access settings.', 'danger')
        return redirect(url_for('login_page'))
        
    user_role = ""
    user_info = {}
    
    if 'admin_id' in session:
        user_role = "admin"
        inst = Institute.query.get(session['admin_id'])
        user_info = {'name': inst.name, 'email': inst.admin_email, 'institute_code': inst.institute_code}
    elif 'teacher_id' in session:
        user_role = "teacher"
        t = Teacher.query.filter_by(teacher_id=session['teacher_id']).first()
        user_info = {'name': t.name, 'email': t.email, 'institute_code': t.institute_code}
    elif 'student_id' in session:
        user_role = "student"
        s = Student.query.get(session['student_id'])
        user_info = {'name': s.name, 'email': s.email, 'institute_code': s.institute_code}

    return render_template('shared/settings.html', user_role=user_role, user_info=user_info)

@app.route('/settings/update_profile', methods=['POST'])
def update_profile():
    if 'admin_id' not in session and 'teacher_id' not in session and 'student_id' not in session:
        return redirect(url_for('login_page'))
    
    new_name = request.form.get('name', '').strip()
    if not new_name:
        flash('Name cannot be empty.', 'danger')
        return redirect(url_for('settings'))

    if 'admin_id' in session:
        inst = Institute.query.get(session['admin_id'])
        if inst:
            inst.name = new_name
            db.session.commit()
    elif 'teacher_id' in session:
        t = Teacher.query.filter_by(teacher_id=session['teacher_id']).first()
        if t:
            t.name = new_name
            db.session.commit()
    elif 'student_id' in session:
        s = Student.query.get(session['student_id'])
        if s:
            s.name = new_name
            db.session.commit()

    flash('Profile updated successfully!', 'success')
    return redirect(url_for('settings'))

@app.route('/settings/change_password', methods=['POST'])
def settings_change_password():
    if 'admin_id' not in session and 'teacher_id' not in session and 'student_id' not in session:
        return redirect(url_for('login_page'))
    
    old_pass = request.form['current_password']
    new_pass = request.form['new_password']
    
    if 'admin_id' in session:
        user = Institute.query.get(session['admin_id'])
        if check_password_hash(user.admin_password, old_pass):
            user.admin_password = generate_password_hash(new_pass)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('settings'))
    elif 'teacher_id' in session:
        user = Teacher.query.filter_by(teacher_id=session['teacher_id']).first()
        if user and check_password_hash(user.password, old_pass):
            user.password = generate_password_hash(new_pass)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('settings'))
    elif 'student_id' in session:
        user = Student.query.get(session['student_id'])
        if user and check_password_hash(user.password, old_pass):
            user.password = generate_password_hash(new_pass)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('settings'))
            
    flash('Incorrect current password.', 'danger')
    return redirect(url_for('settings'))

if __name__ == '__main__':
    app.run(debug=True)
