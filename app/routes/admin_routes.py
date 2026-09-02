from flask import current_app
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from app.models import db
from app.models import *
from app.utils import *
import json
import random
import os
import io
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
import string

from app.routes import main_bp
from app.utils import get_dynamic_time_slots, trim_time_slots, get_val

@main_bp.route('/admin_dash')
def admin_dash():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']
    
    # Calculate Analytics
    c_count = db.session.query(Course.department).filter_by(institute_code=inst_code).distinct().count()
    t_count = Teacher.query.filter_by(institute_code=inst_code).count()
    s_count = Subject.query.filter_by(institute_code=inst_code).count()
    
    # Check if timetable generated
    generated = Timetable.query.filter_by(institute_code=inst_code).first() is not None
    
    # Free teachers (rough logic: those whose assigned hours < max_hours)
    # We can approximate free teachers by checking how many teachers have fewer lectures in Timetable than max_hours, or just list the count of teachers.
    # To keep it efficient:
    all_teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    assigned_counts = {t.name: 0 for t in all_teachers}
    
    if generated:
        # Fix: Count unique slots (day + time) to avoid double counting Common subjects
        tt_entries = Timetable.query.filter_by(institute_code=inst_code).all()
        teacher_slots = {}
        for entry in tt_entries:
            base_name = entry.teacher_name.replace(" (Proxy)", "")
            teacher_slots.setdefault(base_name, set()).add((entry.day_name, entry.start_time))
            
        for name, slots in teacher_slots.items():
            if name in assigned_counts:
                assigned_counts[name] += len(slots)
    
    free_teachers_count = 0
    faculty_workload = []
    
    for t in all_teachers:
        assigned = assigned_counts[t.name]
        free_hrs = max(0, t.max_hours - assigned)
        
        if free_hrs > 0:
            free_teachers_count += 1
            
        faculty_workload.append({
            'name': t.name,
            'max_hours': t.max_hours,
            'assigned_hours': assigned,
            'free_hours': free_hrs
        })

    import math
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    all_courses = Course.query.filter_by(institute_code=inst_code).all()
    
    # Create mapping of class_id to department
    class_dept_map = {c.class_id: c.department for c in all_courses}
    
    weeks_setting = Settings.query.filter_by(institute_code=inst_code, key='weeks_per_semester').first()
    weeks_per_semester = int(weeks_setting.value) if weeks_setting else 15
    
    syllabus_tracking_grouped = {}
    for s in subjects:
        # Attempt to determine department from first class_id mapping
        first_class = s.class_id.split(',')[0].strip() if s.class_id else ""
        dept = class_dept_map.get(first_class, "General/Unassigned")
        
        if dept not in syllabus_tracking_grouped:
            syllabus_tracking_grouped[dept] = []
            
        syllabus_tracking_grouped[dept].append({
            'name': s.subject_name,
            'code': s.subject_code,
            'teacher': s.teacher_id,
            'total_hrs': s.total_course_hours,
            'weekly_hrs': s.required_hours,
            'weeks_needed': math.ceil(s.total_course_hours / s.required_hours) if s.required_hours > 0 else 0
        })

    pending_requests = TeacherUpdateRequest.query.filter_by(institute_code=inst_code, status='Pending').all()
    admin = Institute.query.get(session['admin_id'])
    return render_template('admin/admin_dash.html', admin=admin, c_count=c_count, t_count=t_count, s_count=s_count, generated=generated, free_teachers=free_teachers_count, faculty_workload=faculty_workload, syllabus_tracking_grouped=syllabus_tracking_grouped, weeks_per_semester=weeks_per_semester, pending_requests=pending_requests)

@main_bp.route('/bulk_import/<manage_type>', methods=['POST'])
def bulk_import(manage_type):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']
    
    if 'file' not in request.files:
        flash('No file uploaded.', 'danger')
        return redirect(request.referrer or url_for('main.admin_dash'))
        
    file = request.files['file']
    if file.filename == '':
        flash('No selected file.', 'danger')
        return redirect(request.referrer or url_for('main.admin_dash'))
        
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        flash('Only .csv and .xlsx files are allowed.', 'danger')
        return redirect(request.referrer or url_for('main.admin_dash'))

    success_count = 0
    error_count = 0
    
    try:
        def get_val(r, *keys):
            for k in keys:
                if k in r and r[k] is not None and str(r[k]).strip() != '': return str(r[k]).strip()
            for k in keys:
                target = str(k).lower().replace(' ', '').replace('_', '').replace('/', '')
                for rk in r.keys():
                    if rk is not None:
                        if str(rk).lower().replace(' ', '').replace('_', '').replace('/', '') == target:
                            if r[rk] is not None and str(r[rk]).strip() != '':
                                return str(r[rk]).strip()
            return ''

        data = []
        if file.filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_input = csv.DictReader(stream)
            for row in csv_input:
                data.append(row)
        else:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()))
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data.append(dict(zip(headers, row)))
                    
        for row in data:
            try:
                if manage_type == 'course':
                    class_id = get_val(row, 'class_id', 'classid')
                    dept = get_val(row, 'department', 'departments')
                    sem = get_val(row, 'semester')
                    div = get_val(row, 'division')
                    if not class_id: raise ValueError("class_id missing")
                    c = Course(institute_code=inst_code, class_id=class_id, department=dept, semester=sem, division=div)
                    db.session.add(c)
                elif manage_type == 'teacher':
                    tid = get_val(row, 'teacher_id', 'teacherid')
                    name = get_val(row, 'name')
                    email = get_val(row, 'email')
                    depts = get_val(row, 'departments', 'department')
                    hours = get_val(row, 'max_hours', 'max_hours_week', 'maxhoursweek', 'maxhours')
                    days = get_val(row, 'available_days', 'days')
                    if not tid or not name: raise ValueError("teacher_id or name missing")
                    t = Teacher(institute_code=inst_code, teacher_id=tid, name=name, email=email, departments=depts, max_hours=int(hours or 0), available_days=days)
                    db.session.add(t)
                elif manage_type == 'subject':
                    scode = get_val(row, 'subject_code', 'subjectcode')
                    sname = get_val(row, 'subject_name', 'subjectname', 'subject')
                    cid = get_val(row, 'class_id', 'classid')
                    tid = get_val(row, 'teacher_id', 'teacherid')
                    stype = get_val(row, 'subject_type', 'subjecttype') or 'Theory'
                    req_hrs_raw = get_val(row, 'required_hours', 'requiredhours')
                    tot_hrs = get_val(row, 'total_course_hours', 'totalcoursehours', 'totalhours') or 50
                    sess_len = get_val(row, 'session_length', 'sessionlength') or 1
                    
                    if req_hrs_raw:
                        req_hrs = int(req_hrs_raw)
                    else:
                        weeks_setting = Settings.query.filter_by(institute_code=inst_code, key='weeks_per_semester').first()
                        weeks = int(weeks_setting.value) if weeks_setting else 15
                        import math
                        req_hrs = math.ceil(int(tot_hrs) / (weeks * int(sess_len)))
                        
                    if not scode: raise ValueError("subject_code missing")
                    s = Subject(institute_code=inst_code, subject_code=scode, subject_name=sname, class_id=cid, teacher_id=tid, subject_type=stype, required_hours=req_hrs, total_course_hours=int(tot_hrs), session_length=int(sess_len))
                    db.session.add(s)
                db.session.commit()
                success_count += 1
            except Exception as e:
                db.session.rollback()
                print(f"Error importing row {row}: {e}")
                error_count += 1
                
        flash(f'Batch Upload Complete! Successfully added: {success_count}. Failed/Duplicates: {error_count}.', 'info')
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'danger')
        
    if manage_type == 'course': return redirect(url_for('main.manage_courses'))
    elif manage_type == 'teacher': return redirect(url_for('main.manage_teachers'))
    elif manage_type == 'subject': return redirect(url_for('main.manage_subjects'))
    return redirect(url_for('main.admin_dash'))

@main_bp.route('/manage_courses', methods=['GET', 'POST'])
def manage_courses():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        db.session.add(Course(
            institute_code=inst_code, class_id=request.form['class_id'],
            department=request.form['department'], semester=request.form['semester'],
            division=request.form['division']
        ))
        db.session.commit()
        flash('Course added!', 'success')
        return redirect(url_for('main.manage_courses'))
    
    courses = Course.query.filter_by(institute_code=inst_code).all()
    return render_template('admin/manage_master.html', manage_type='course', items=courses)

@main_bp.route('/manage_teachers', methods=['GET', 'POST'])
def manage_teachers():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        if Teacher.query.filter_by(email=request.form['email']).first():
            flash('Email already registered!', 'danger')
            return redirect(url_for('main.manage_teachers'))
            
        days = request.form.getlist('days')
        if not days:
            flash('Select at least one day!', 'danger')
            return redirect(url_for('main.manage_teachers'))

        db.session.add(Teacher(
            institute_code=inst_code, teacher_id=request.form['teacher_id'],
            name=request.form['name'], email=request.form['email'],
            departments=request.form['departments'], available_days=",".join(days),
            max_hours=int(request.form['max_hours'])
        ))
        db.session.commit()
        flash('Teacher added!', 'success')
        return redirect(url_for('main.manage_teachers'))
    
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    unique_depts = list(set([c.department for c in courses]))
    return render_template('admin/manage_master.html', manage_type='teacher', items=teachers, depts=unique_depts)

@main_bp.route('/manage_subjects', methods=['GET', 'POST'])
def manage_subjects():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        class_ids = request.form.getlist('class_id')
        if not class_ids:
            flash('Select at least one Class!', 'danger')
            return redirect(url_for('main.manage_subjects'))

        weeks_setting = Settings.query.filter_by(institute_code=inst_code, key='weeks_per_semester').first()
        weeks = int(weeks_setting.value) if weeks_setting else 15
        total_hours = int(request.form['total_course_hours'])
        session_len = int(request.form.get('session_length', 1))
        import math
        
        db.session.add(Subject(
            institute_code=inst_code, subject_code=request.form['subject_code'],
            subject_name=request.form['subject_name'], class_id=",".join(class_ids),
            teacher_id=request.form['teacher_id'], total_course_hours=total_hours,
            required_hours=math.ceil(total_hours / (weeks * session_len)),
            subject_type=request.form['subject_type'], session_length=session_len
        ))
        db.session.commit()
        flash('Subject mapped!', 'success')
        return redirect(url_for('main.manage_subjects'))
    
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    return render_template('admin/manage_master.html', manage_type='subject', items=subjects, courses=courses, teachers=teachers)

@main_bp.route('/edit_course/<int:id>', methods=['GET', 'POST'])
def edit_course(id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    course = Course.query.get_or_404(id)
    if request.method == 'POST':
        course.class_id = request.form['class_id']
        course.department = request.form['department']
        course.semester = request.form['semester']
        course.division = request.form['division']
        db.session.commit()
        flash('Course updated!', 'success')
        return redirect(url_for('main.manage_courses'))
    return render_template('admin/edit_master.html', item=course, edit_type='course')

@main_bp.route('/edit_teacher/<int:id>', methods=['GET', 'POST'])
def edit_teacher(id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    teacher = Teacher.query.get_or_404(id)
    if request.method == 'POST':
        teacher.teacher_id = request.form['teacher_id']
        teacher.name = request.form['name']
        teacher.email = request.form['email']
        teacher.departments = request.form['departments']
        teacher.max_hours = request.form['max_hours']
        days = request.form.getlist('days')
        if days: teacher.available_days = ",".join(days)
        db.session.commit()
        flash('Teacher updated!', 'success')
        return redirect(url_for('main.manage_teachers'))
    
    courses = Course.query.filter_by(institute_code=session['institute_code']).all()
    depts = list(set([c.department for c in courses]))
    return render_template('admin/edit_master.html', item=teacher, edit_type='teacher', unique_depts=depts)

@main_bp.route('/edit_subject/<int:id>', methods=['GET', 'POST'])
def edit_subject(id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    subject = Subject.query.get_or_404(id)
    if request.method == 'POST':
        subject.subject_code = request.form['subject_code']
        subject.subject_name = request.form['subject_name']
        class_ids = request.form.getlist('class_id')
        if class_ids: subject.class_id = ",".join(class_ids)
        subject.teacher_id = request.form['teacher_id']
        weeks_setting = Settings.query.filter_by(institute_code=session['institute_code'], key='weeks_per_semester').first()
        weeks = int(weeks_setting.value) if weeks_setting else 15
        total_hours = int(request.form['total_course_hours'])
        session_len = int(request.form.get('session_length', 1))
        import math
        subject.total_course_hours = total_hours
        subject.session_length = session_len
        subject.required_hours = math.ceil(total_hours / (weeks * session_len))
        subject.subject_type = request.form['subject_type']
        db.session.commit()
        flash('Subject updated!', 'success')
        return redirect(url_for('main.manage_subjects'))
    
    courses = Course.query.filter_by(institute_code=session['institute_code']).all()
    teachers = Teacher.query.filter_by(institute_code=session['institute_code']).all()
    return render_template('admin/edit_master.html', item=subject, edit_type='subject', courses=courses, teachers=teachers)

@main_bp.route('/delete/<type>/<int:id>')
def delete_item(type, id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    
    if type == 'course':
        item = Course.query.get_or_404(id)
        route = 'manage_courses'
    elif type == 'teacher':
        item = Teacher.query.get_or_404(id)
        route = 'manage_teachers'
    else:
        item = Subject.query.get_or_404(id)
        route = 'manage_subjects'

    if item.institute_code == session['institute_code']:
        db.session.delete(item)
        db.session.commit()
        flash('Deleted successfully.', 'info')
    
    return redirect(url_for('main.' + route))

@main_bp.route('/bulk_delete/<type>', methods=['POST'])
def bulk_delete_items(type):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']
    selected_ids = request.form.getlist('selected_ids')
    
    if not selected_ids:
        flash('No items selected for deletion.', 'warning')
        if type == 'course': return redirect(url_for('main.manage_courses'))
        if type == 'teacher': return redirect(url_for('main.manage_teachers'))
        return redirect(url_for('main.manage_subjects'))
    
    try:
        if type == 'course':
            num_deleted = Course.query.filter(Course.id.in_(selected_ids), Course.institute_code==inst_code).delete(synchronize_session=False)
            route = 'manage_courses'
        elif type == 'teacher':
            num_deleted = Teacher.query.filter(Teacher.id.in_(selected_ids), Teacher.institute_code==inst_code).delete(synchronize_session=False)
            route = 'manage_teachers'
        elif type == 'subject':
            num_deleted = Subject.query.filter(Subject.id.in_(selected_ids), Subject.institute_code==inst_code).delete(synchronize_session=False)
            route = 'manage_subjects'
            
        db.session.commit()
        flash(f'Successfully deleted {num_deleted} items.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting items: {str(e)}', 'danger')
        route = 'admin_dash'
        
    return redirect(url_for('main.' + route))

@main_bp.route('/generate_timetable', methods=['GET', 'POST'])
def generate_timetable():
    if request.method == 'GET':
        return redirect(url_for('main.admin_dash'))
        
    if 'admin_id' not in session:
        return redirect(url_for('main.login_page'))
        
    inst_code = session['institute_code']
    
    # 🧹 1. Clear existing timetable for this institute (Fresh Start)
    Timetable.query.filter_by(institute_code=inst_code).delete()
    db.session.commit()
    
    # 📥 2. Fetch all required Data
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    
    # Dictionary for quick Teacher lookups
    teacher_dict = {t.teacher_id: t for t in teachers}
    
    # Get Time Slots (Assuming your get_dynamic_time_slots function is present)
    time_slots = get_dynamic_time_slots(inst_code) 
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    
    # 🧠 3. Initialize Tracking States (To prevent Clashes)
    # Track which class is busy at what time
    class_timetable = {c.class_id: {day: {} for day in days} for c in courses}
    # Track which teacher is busy at what time
    teacher_timetable = {t.teacher_id: {day: {} for day in days} for t in teachers}
    # Track maximum hours given to a teacher
    teacher_hours = {t.teacher_id: 0 for t in teachers}
    
    # ⚔️ Phase Separation
    common_subjects = []
    normal_subjects = []
    
    for sub in subjects:
        if ',' in sub.class_id:
            common_subjects.append(sub)  # e.g. "FYCS, FYIT, FYBMS"
        else:
            normal_subjects.append(sub)

    # 🚀 PHASE 1: Process Common Subjects (Highest Priority)
    import random
    random.shuffle(common_subjects)
    for sub in common_subjects:
        # Split and clean class IDs
        target_classes = [c.strip() for c in sub.class_id.split(',')]
        assigned_hours = 0
        teacher = teacher_dict.get(sub.teacher_id)
        
        while assigned_hours < sub.required_hours:
            scheduled_this_round = False
            # Randomness + Balanced Load
            random.shuffle(days)
            days.sort(key=lambda d: max(len(class_timetable.get(c, {}).get(d, {})) for c in target_classes))
            
            for day in days:
                if assigned_hours >= sub.required_hours: break
                
                # Constraint: Preferred Days for Subject
                if sub.preferred_days and day not in sub.preferred_days: continue
                
                # Constraint: Is teacher available on this day?
                if teacher and day not in teacher.available_days: continue
                
                # Check for contiguous blocks based on session_length
                valid_indices = list(range(len(time_slots) - sub.session_length + 1))
                
                for idx in valid_indices:
                    slots_to_check = time_slots[idx : idx + sub.session_length]
                    
                    classes_free = all(s[0] not in class_timetable.get(c, {}).get(day, {}) for c in target_classes for s in slots_to_check)
                    teacher_free = all(s[0] not in teacher_timetable.get(sub.teacher_id, {}).get(day, {}) for s in slots_to_check)
                    hours_ok = teacher_hours.get(sub.teacher_id, 0) + sub.session_length <= teacher.max_hours if teacher else True
                    
                    if classes_free and teacher_free and hours_ok:
                        # ✅ ASSIGN TO ALL CLASSES SIMULTANEOUSLY
                        for s in slots_to_check:
                            for c in target_classes:
                                if c in class_timetable:
                                    class_timetable[c][day][s[0]] = (sub, s[1])
                            teacher_timetable[sub.teacher_id][day][s[0]] = (sub, s[1])
                            
                        if teacher: teacher_hours[sub.teacher_id] += sub.session_length
                        assigned_hours += sub.session_length
                        scheduled_this_round = True
                        break # Move to next day (Horizontal distribution)
                        
            if not scheduled_this_round: break # Stuck

    # 🚀 PHASE 2: Process Normal Subjects
    random.shuffle(normal_subjects)
    for sub in normal_subjects:
        assigned_hours = 0
        target_class = sub.class_id
        teacher = teacher_dict.get(sub.teacher_id)
        
        if target_class not in class_timetable: continue
        
        while assigned_hours < sub.required_hours:
            scheduled_this_round = False
            # Randomness + Balanced Load
            random.shuffle(days)
            days.sort(key=lambda d: len(class_timetable[target_class][d]))
            
            for day in days:
                if assigned_hours >= sub.required_hours: break
                
                # Constraint: Preferred Days for Subject
                if sub.preferred_days and day not in sub.preferred_days: continue
                
                # Constraint: Is teacher available on this day?
                if teacher and day not in teacher.available_days: continue
                
                valid_indices = list(range(len(time_slots) - sub.session_length + 1))
                
                for idx in valid_indices:
                    slots_to_check = time_slots[idx : idx + sub.session_length]
                    
                    # CLASH DETECTION
                    class_free = all(s[0] not in class_timetable[target_class][day] for s in slots_to_check)
                    teacher_free = all(s[0] not in teacher_timetable.get(sub.teacher_id, {}).get(day, {}) for s in slots_to_check)
                    hours_ok = teacher_hours.get(sub.teacher_id, 0) + sub.session_length <= teacher.max_hours if teacher else True
                    
                    if class_free and teacher_free and hours_ok:
                        # ✅ ASSIGN LECTURE
                        for s in slots_to_check:
                            class_timetable[target_class][day][s[0]] = (sub, s[1])
                            teacher_timetable[sub.teacher_id][day][s[0]] = (sub, s[1])
                            
                        if teacher: teacher_hours[sub.teacher_id] += sub.session_length
                        assigned_hours += sub.session_length
                        scheduled_this_round = True
                        break # Move to next day (Horizontal distribution)
                        
            if not scheduled_this_round: break # Stuck
    # Removed Phase 3 (Aggressive Compaction) because it broke session_length blocks.

    # 🚀 PHASE 4: Cross-Day Gap Elimination (Move isolated lectures to other days)
    for c_id in class_timetable.keys():
        for _ in range(3): # Multiple passes to resolve cascading gaps
            for day in days:
                slots_data = class_timetable[c_id].get(day, {})
                if not slots_data: continue
                
                filled_indices = [i for i, slot in enumerate(time_slots) if slot[0] in slots_data]
                if not filled_indices: continue
                
                if max(filled_indices) >= len(filled_indices):
                    # Gap detected! Find the lectures placed AFTER the gap
                    sorted_filled = sorted(filled_indices)
                    for i, idx in enumerate(sorted_filled):
                        if idx > i: # This lecture is separated by a gap
                            st_time = time_slots[idx][0]
                            sub_data = class_timetable[c_id][day][st_time]
                            sub = sub_data[0]
                            
                            # Do not attempt to move common subjects or block subjects (session_length > 1)
                            if ',' in sub.class_id or sub.session_length > 1: continue
                            
                            moved = False
                            for other_day in days:
                                if other_day == day: continue
                                
                                # Find first empty slot on other_day
                                other_slots_data = class_timetable[c_id].get(other_day, {})
                                target_st, target_end = None, None
                                for slot in time_slots:
                                    if slot[0] not in other_slots_data:
                                        target_st, target_end = slot[0], slot[1]
                                        break
                                        
                                if target_st and target_st not in teacher_timetable.get(sub.teacher_id, {}).get(other_day, {}):
                                    # Move successful!
                                    del class_timetable[c_id][day][st_time]
                                    del teacher_timetable[sub.teacher_id][day][st_time]
                                    
                                    class_timetable[c_id][other_day][target_st] = (sub, target_end)
                                    teacher_timetable.setdefault(sub.teacher_id, {}).setdefault(other_day, {})[target_st] = (sub, target_end)
                                    moved = True
                                    break
                            
                            if moved:
                                break # Restart pass for this class since timetable mutated

    # 💾 4. Save the Final Generated Output to Database
    records_to_add = []
    for c_id, days_data in class_timetable.items():
        for day, slots_data in days_data.items():
            for start_time, sub_data in slots_data.items():
                
                sub = sub_data[0]
                end_time = sub_data[1]
                
                teacher = teacher_dict.get(sub.teacher_id)
                t_name = teacher.name if teacher else sub.teacher_id
                
                # Create Database Entry
                new_entry = Timetable(
                    institute_code=inst_code,
                    class_id=c_id,
                    day_name=day,
                    start_time=start_time,
                    end_time=end_time,
                    subject_name=sub.subject_name,
                    teacher_name=t_name,
                    is_proxy=False  # Proxy is always False during master generation
                )
                records_to_add.append(new_entry)
                
    db.session.bulk_save_objects(records_to_add)
    db.session.commit()
    
    flash('⚡ Timetable Generated Successfully! Zero Clashes Detected.', 'success')
    return redirect(url_for('main.admin_dash'))

@main_bp.route('/view_timetable')
def view_timetable():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']
    
    courses = Course.query.filter_by(institute_code=inst_code).all()
    selected_class = request.args.get('class_id')
    
    schedule = {}
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    time_slots = get_dynamic_time_slots(inst_code)
    
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    lunch_after = int(s.get('lunch_after_lecture', 2))
    break_duration = int(s.get('break_time', 30)) # Fetch break duration
    
    if selected_class:
        entries = Timetable.query.filter_by(institute_code=inst_code, class_id=selected_class).all()
        for day in days:
            schedule[day] = {}
            for entry in entries:
                if entry.day_name == day:
                    schedule[day][entry.start_time] = entry
        time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    inst = Institute.query.filter_by(institute_code=inst_code).first()
    inst_name = inst.name if inst else "Institute"
    
    # Fetch teachers and subjects for manual editing
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    if selected_class:
        subjects = Subject.query.filter_by(institute_code=inst_code, class_id=selected_class).all()
    else:
        subjects = []
    
    return render_template('shared/view_timetable.html', courses=courses, selected_class=selected_class, schedule=schedule, days=days, time_slots=time_slots, lunch_after=lunch_after, break_duration=break_duration, inst_name=inst_name, teachers=teachers, subjects=subjects)

@main_bp.route('/edit_timetable_slot', methods=['POST'])
def edit_timetable_slot():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    
    entry_id = request.form.get('entry_id')
    new_subject = request.form.get('new_subject')
    new_teacher = request.form.get('new_teacher')
    
    entry = Timetable.query.get(entry_id)
    if not entry:
        flash('Timetable slot not found.', 'danger')
        return redirect(url_for('main.view_timetable', class_id=request.form.get('class_id')))
        
    # Check for teacher clash (is the new teacher busy at this exact day and time?)
    if new_teacher != entry.teacher_name:
        clash = Timetable.query.filter_by(
            institute_code=entry.institute_code,
            day_name=entry.day_name,
            start_time=entry.start_time,
            teacher_name=new_teacher
        ).first()
        
        if clash:
            flash(f'Cannot assign {new_teacher}. They are already teaching Class {clash.class_id} at this time!', 'danger')
            return redirect(url_for('main.view_timetable', class_id=entry.class_id))
            
    # Update entry
    entry.subject_name = new_subject
    entry.teacher_name = new_teacher
    # Clear proxy flag if any
    if hasattr(entry, 'is_proxy') and entry.is_proxy:
        entry.is_proxy = False
        entry.original_teacher = None
        
    db.session.commit()
    flash('Timetable slot updated successfully!', 'success')
    return redirect(url_for('main.view_timetable', class_id=entry.class_id))

@main_bp.route('/export_timetables')
def export_timetables():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']
    
    # Fetch Classes
    classes = db.session.query(Timetable.class_id).filter_by(institute_code=inst_code).distinct().all()
    class_ids = [c[0] for c in classes]
    
    if not class_ids:
        flash('No timetable generated yet to export!', 'warning')
        return redirect(url_for('main.view_timetable'))

    # Fetch Settings & Slots
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}
    
    institute_name = s.get('institute_name', 'INSTITUTE TIMETABLE').upper()
    lunch_after = int(s.get('lunch_after_lecture', 2))
    
    # Time slots format: "08:00 AM - 09:00 AM"
    time_slots = get_dynamic_time_slots(inst_code)
    slots = [f"{slot[0]} - {slot[1]}" for slot in time_slots]
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

    # Excel Styling Variables
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    cell_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    lunch_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cls in class_ids:
        safe_sheet_name = str(cls).replace('/', '-').replace('\\', '-')[:31]
        ws = wb.create_sheet(title=safe_sheet_name)
        
        # Titles
        c1 = ws.cell(row=1, column=1, value=institute_name)
        c1.font = title_font; c1.alignment = center_align
        c2 = ws.cell(row=2, column=1, value=f"TIMETABLE / SCHEDULE - {cls}")
        c2.font = header_font; c2.alignment = center_align; c2.fill = blue_fill
        
        ws.cell(row=3, column=1, value="Day").font = header_font
        ws.cell(row=3, column=1).alignment = center_align; ws.cell(row=3, column=1).border = thin_border
        ws.column_dimensions['A'].width = 12
        
        col_idx = 2
        lunch_printed = False
        slot_to_col = {}
        
        # Build Header & Lunch Break Column
        for i, slot in enumerate(slots):
            # Print Lunch Column exactly after the specified lecture
            if not lunch_printed and i == lunch_after:
                lc = ws.cell(row=3, column=col_idx, value="LUNCH\nBREAK")
                lc.font = header_font; lc.alignment = center_align; lc.border = thin_border; lc.fill = lunch_fill
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10
                
                # Merge and rotate text vertically
                ws.merge_cells(start_row=4, start_column=col_idx, end_row=4+len(days)-1, end_column=col_idx)
                lb_cell = ws.cell(row=4, column=col_idx, value="L U N C H   B R E A K")
                lb_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center', wrap_text=True)
                lb_cell.font = Font(name='Arial', size=14, bold=True, color="595959")
                lb_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                for r in range(4, 4+len(days)): ws.cell(row=r, column=col_idx).border = thin_border
                col_idx += 1
                lunch_printed = True
            
            # Print Time Slot Header
            slot_to_col[slot] = col_idx
            cell = ws.cell(row=3, column=col_idx, value=slot.replace(" - ", "\n"))
            cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
            col_idx += 1
            
        max_col = col_idx - 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            
        # Fetch data for this specific class
        tt_entries = Timetable.query.filter_by(institute_code=inst_code, class_id=cls).all()
        tt_dict = {(e.day_name, f"{e.start_time} - {e.end_time}"): f"{e.subject_name}\n({e.teacher_name})" for e in tt_entries}
        
        # Populate Grid Data
        for row_idx, day in enumerate(days, start=4):
            day_cell = ws.cell(row=row_idx, column=1, value=day)
            day_cell.font = header_font; day_cell.alignment = center_align; day_cell.border = thin_border
            
            for slot in slots:
                c_idx = slot_to_col[slot]
                cell = ws.cell(row=row_idx, column=c_idx, value=tt_dict.get((day, slot), "---"))
                cell.font = cell_font; cell.alignment = center_align; cell.border = thin_border
                
    # Save Excel to memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="AutoTime_Master_Schedule.xlsx", as_attachment=True)

@main_bp.route('/college_settings', methods=['GET', 'POST'])
def college_settings():
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    inst_code = session['institute_code']

    if request.method == 'POST':
        # Yahan humne 'lunch_after_lecture' add kar diya hai
        keys = ['total_lectures', 'start_time', 'lecture_duration', 'break_time', 'lunch_after_lecture', 'weeks_per_semester']
        for key in keys:
            val = request.form.get(key)
            if val:
                setting = Settings.query.filter_by(institute_code=inst_code, key=key).first()
                if setting:
                    setting.value = val 
                else:
                    new_setting = Settings(institute_code=inst_code, key=key, value=val)
                    db.session.add(new_setting)
        
        db.session.commit()
        flash('College Settings Updated Successfully!', 'success')
        return redirect(url_for('main.college_settings'))

    all_settings = Settings.query.filter_by(institute_code=inst_code).all()
    settings_dict = {s.key: s.value for s in all_settings}

    return render_template('admin/college_settings.html', settings=settings_dict)

@main_bp.route('/admin/requests/approve/<int:req_id>', methods=['POST'])
def approve_teacher_request(req_id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    req = TeacherUpdateRequest.query.get_or_404(req_id)
    if req.institute_code != session.get('institute_code'): return "Unauthorized", 403
    
    teacher = Teacher.query.filter_by(teacher_id=req.teacher_id, institute_code=req.institute_code).first()
    if teacher:
        if req.new_name:
            teacher.name = req.new_name
        if req.new_email:
            teacher.email = req.new_email
        req.status = 'Approved'
        db.session.commit()
        flash('Teacher update request approved.', 'success')
    return redirect(url_for('main.admin_dash'))

@main_bp.route('/admin/requests/reject/<int:req_id>', methods=['POST'])
def reject_teacher_request(req_id):
    if 'admin_id' not in session: return redirect(url_for('main.login_page'))
    req = TeacherUpdateRequest.query.get_or_404(req_id)
    if req.institute_code != session.get('institute_code'): return "Unauthorized", 403
    
    req.status = 'Rejected'
    db.session.commit()
    flash('Teacher update request rejected.', 'info')
    return redirect(url_for('main.admin_dash'))