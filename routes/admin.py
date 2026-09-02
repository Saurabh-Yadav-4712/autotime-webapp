from utils.timetable_helpers import build_timetable_view_model
from utils.decorators import login_required_admin
from flask import (
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from models import (
    db,
    Institute,
    Course,
    Subject,
    Teacher,
    Timetable,
    Settings,
    Student,
    TeacherUpdateRequest,
    AcademicCalendar,
    TeacherLeave,
    Notification,
    GenerationHistory,
)
from utils.helpers import (
    get_dynamic_time_slots,
    get_val,
    is_valid_email,
    normalize_email,
    trim_time_slots,
)
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
from io import BytesIO
from datetime import datetime

from routes.blueprint import main_bp


def _tenant_record_or_404(model, record_id):
    return model.query.filter_by(
        id=record_id,
        institute_code=session["institute_code"],
    ).first_or_404()


def _parse_form_int(field_name, minimum, maximum, default=None):
    raw_value = request.form.get(field_name, default)
    try:
        value = int(raw_value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name.replace('_', ' ').title()} must be a number.") from exc
    if value < minimum or value > maximum:
        raise ValueError(
            f"{field_name.replace('_', ' ').title()} must be between {minimum} and {maximum}."
        )
    return value


@main_bp.route("/admin_dash")
@login_required_admin
def admin_dash():
    inst_code = session["institute_code"]

    # Calculate Analytics
    c_count = (
        db.session.query(Course.department).filter_by(institute_code=inst_code).distinct().count()
    )
    t_count = Teacher.query.filter_by(institute_code=inst_code).count()
    s_count = Subject.query.filter_by(institute_code=inst_code).count()

    # Check if timetable generated
    generated = Timetable.query.filter_by(institute_code=inst_code).first() is not None

    # Free teachers (rough logic: those whose assigned hours < max_hours)
    # We can approximate free teachers by checking how many teachers have fewer lectures in Timetable than max_hours, or just list the count of teachers.
    # To keep it efficient:
    all_teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    assigned_counts = {t.name: 0 for t in all_teachers}

    if generated:
        # Fix: Count unique slots (day + time) to avoid double counting Common subjects
        tt_entries = Timetable.query.filter_by(institute_code=inst_code).all()
        teacher_slots = {}
        for entry in tt_entries:
            base_name = entry.teacher_name.replace(" (Proxy)", "")
            teacher_slots.setdefault(base_name, set()).add((entry.day_name, entry.start_time))

        for name, slots in teacher_slots.items():
            if name in assigned_counts:
                assigned_counts[name] += len(slots)

    free_teachers_count = 0
    faculty_workload = []

    for t in all_teachers:
        assigned = assigned_counts[t.name]
        free_hrs = max(0, t.max_hours - assigned)

        if free_hrs > 0:
            free_teachers_count += 1

        faculty_workload.append(
            {
                "name": t.name,
                "max_hours": t.max_hours,
                "assigned_hours": assigned,
                "free_hours": free_hrs,
            }
        )

    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    all_courses = Course.query.filter_by(institute_code=inst_code).all()

    pending_requests = TeacherUpdateRequest.query.filter_by(
        institute_code=inst_code, status="Pending"
    ).all()
    admin = db.session.get(Institute, session["admin_id"])
    return render_template(
        "admin/admin_dash.html",
        admin=admin,
        c_count=c_count,
        t_count=t_count,
        s_count=s_count,
        generated=generated,
        free_teachers=free_teachers_count,
        faculty_workload=faculty_workload,
        pending_requests=pending_requests,
    )


@main_bp.route("/bulk_import/<manage_type>", methods=["POST"])
@login_required_admin
def bulk_import(manage_type):
    inst_code = session["institute_code"]
    if manage_type not in {"course", "teacher", "subject"}:
        return "Unsupported import type", 404

    if "file" not in request.files:
        flash("No file uploaded.", "danger")
        return redirect(request.referrer or url_for("main.admin_dash"))

    file = request.files["file"]
    if file.filename == "":
        flash("No selected file.", "danger")
        return redirect(request.referrer or url_for("main.admin_dash"))

    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        flash("Only .csv and .xlsx files are allowed.", "danger")
        return redirect(request.referrer or url_for("main.admin_dash"))

    success_count = 0
    error_count = 0

    try:

        def get_val(r, *keys):
            for k in keys:
                if k in r and r[k] is not None and str(r[k]).strip() != "":
                    return str(r[k]).strip()
            for k in keys:
                target = str(k).lower().replace(" ", "").replace("_", "").replace("/", "")
                for rk in r.keys():
                    if rk is not None:
                        if (
                            str(rk).lower().replace(" ", "").replace("_", "").replace("/", "")
                            == target
                        ):
                            if r[rk] is not None and str(r[rk]).strip() != "":
                                return str(r[rk]).strip()
            return ""

        data = []
        if filename.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            csv_input = csv.DictReader(stream)
            for row in csv_input:
                data.append(row)
        else:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(file.read()),
                read_only=True,
                data_only=True,
            )
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data.append(dict(zip(headers, row)))

        for row in data:
            try:
                if manage_type == "course":
                    class_id = get_val(row, "class_id", "classid")
                    dept = get_val(row, "department", "departments")
                    sem = get_val(row, "semester")
                    div = get_val(row, "division")
                    if not class_id:
                        raise ValueError("class_id missing")
                    c = Course(
                        institute_id=session.get("admin_id"),
                        institute_code=inst_code,
                        class_id=class_id,
                        department=dept,
                        semester=sem,
                        division=div,
                    )
                    db.session.add(c)
                elif manage_type == "teacher":
                    tid = get_val(row, "teacher_id", "teacherid")
                    name = get_val(row, "name")
                    email = get_val(row, "email")
                    depts = get_val(row, "departments", "department")
                    hours = get_val(row, "max_hours", "max_hours_week", "maxhoursweek", "maxhours")
                    days = get_val(row, "available_days", "days")
                    if not tid or not name:
                        raise ValueError("teacher_id or name missing")
                    t = Teacher(
                        institute_id=session.get("admin_id"),
                        institute_code=inst_code,
                        teacher_id=tid,
                        name=name,
                        email=email,
                        departments=depts,
                        max_hours=int(hours or 0),
                        available_days=days,
                    )
                    db.session.add(t)
                elif manage_type == "subject":
                    scode = get_val(row, "subject_code", "subjectcode")
                    sname = get_val(row, "subject_name", "subjectname", "subject")
                    cid = get_val(row, "class_id", "classid")
                    tid = get_val(row, "teacher_id", "teacherid")
                    stype = get_val(row, "subject_type", "subjecttype") or "Theory"
                    req_hrs_raw = get_val(row, "required_hours", "requiredhours")
                    tot_hrs = (
                        get_val(row, "total_course_hours", "totalcoursehours", "totalhours") or 50
                    )
                    sess_len = get_val(row, "session_length", "sessionlength") or 1

                    req_hrs = int(req_hrs_raw) if req_hrs_raw else 4
                    sess_len_int = int(sess_len)
                    if not scode:
                        raise ValueError("subject_code missing")
                    if req_hrs % sess_len_int != 0:
                        raise ValueError("required_hours not divisible by session_length")

                    t_db = Teacher.query.filter_by(institute_code=inst_code, teacher_id=tid).first()
                    t_fk = t_db.id if t_db else None

                    s = Subject(
                        institute_id=session.get("admin_id"),
                        institute_code=inst_code,
                        subject_code=scode,
                        subject_name=sname,
                        class_id=cid,
                        teacher_id=tid,
                        teacher_id_fk=t_fk,
                        subject_type=stype,
                        required_hours=req_hrs,
                        total_course_hours=int(tot_hrs),
                        session_length=sess_len_int,
                    )

                    if cid:
                        c_list = [c.strip() for c in cid.split(",")]
                        courses_db = Course.query.filter(Course.institute_code == inst_code, Course.class_id.in_(c_list)).all()
                        for c_obj in courses_db:
                            s.courses.append(c_obj)

                    db.session.add(s)
                db.session.commit()
                success_count += 1
            except Exception:
                db.session.rollback()
                current_app.logger.exception("Bulk import row failed")
                error_count += 1

        flash(
            f"Batch Upload Complete! Successfully added: {success_count}. Failed/Duplicates: {error_count}.",
            "info",
        )
    except Exception:
        current_app.logger.exception("Bulk import failed")
        flash("The uploaded file could not be processed. Check its format and values.", "danger")

    if manage_type == "course":
        return redirect(url_for("main.manage_courses"))
    elif manage_type == "teacher":
        return redirect(url_for("main.manage_teachers"))
    elif manage_type == "subject":
        return redirect(url_for("main.manage_subjects"))
    return redirect(url_for("main.admin_dash"))


@main_bp.route("/manage_courses", methods=["GET", "POST"])
@login_required_admin
def manage_courses():
    inst_code = session["institute_code"]

    if request.method == "POST":
        class_id = request.form.get("class_id", "").strip()
        department = request.form.get("department", "").strip()
        division = request.form.get("division", "").strip()
        try:
            semester = _parse_form_int("semester", 1, 20)
        except ValueError as error:
            flash(str(error), "danger")
            return redirect(url_for("main.manage_courses"))
        if not class_id or not department or not division:
            flash("Class ID, department, and division are required.", "danger")
            return redirect(url_for("main.manage_courses"))
        if Course.query.filter_by(institute_code=inst_code, class_id=class_id).first():
            flash("That class ID already exists.", "danger")
            return redirect(url_for("main.manage_courses"))
        db.session.add(
            Course(
                institute_id=session.get("admin_id"),
                institute_code=inst_code,
                class_id=class_id,
                department=department,
                semester=semester,
                division=division,
            )
        )
        db.session.commit()
        flash("Course added!", "success")
        return redirect(url_for("main.manage_courses"))

    courses = Course.query.filter_by(institute_code=inst_code).all()
    return render_template("admin/manage_master.html", manage_type="course", items=courses)


@main_bp.route("/manage_teachers", methods=["GET", "POST"])
@login_required_admin
def manage_teachers():
    inst_code = session["institute_code"]

    if request.method == "POST":
        teacher_id = request.form.get("teacher_id", "").strip()
        name = request.form.get("name", "").strip()
        email = normalize_email(request.form.get("email"))
        departments = request.form.get("departments", "").strip()
        if not teacher_id or not name or not departments or not is_valid_email(email):
            flash("Teacher ID, name, department, and a valid email are required.", "danger")
            return redirect(url_for("main.manage_teachers"))
        if (
            Teacher.query.filter_by(email=email).first()
            or Institute.query.filter_by(admin_email=email).first()
            or Student.query.filter_by(email=email).first()
        ):
            flash("Email already registered!", "danger")
            return redirect(url_for("main.manage_teachers"))
        if Teacher.query.filter_by(institute_code=inst_code, teacher_id=teacher_id).first():
            flash("That teacher ID already exists in this institute.", "danger")
            return redirect(url_for("main.manage_teachers"))

        days = request.form.getlist("days")
        if not days:
            flash("Select at least one day!", "danger")
            return redirect(url_for("main.manage_teachers"))

        try:
            max_hours = _parse_form_int("max_hours", 1, 100)
        except ValueError as error:
            flash(str(error), "danger")
            return redirect(url_for("main.manage_teachers"))
        db.session.add(
            Teacher(
                institute_id=session.get("admin_id"),
                institute_code=inst_code,
                teacher_id=teacher_id,
                name=name,
                email=email,
                departments=departments,
                available_days=",".join(days),
                max_hours=max_hours,
            )
        )
        db.session.commit()
        flash("Teacher added!", "success")
        return redirect(url_for("main.manage_teachers"))

    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    unique_depts = list(set([c.department for c in courses]))
    return render_template(
        "admin/manage_master.html", manage_type="teacher", items=teachers, depts=unique_depts
    )


@main_bp.route("/manage_subjects", methods=["GET", "POST"])
@login_required_admin
def manage_subjects():
    inst_code = session["institute_code"]

    if request.method == "POST":
        class_ids = request.form.getlist("class_id")
        if not class_ids:
            flash("Select at least one Class!", "danger")
            return redirect(url_for("main.manage_subjects"))

        valid_class_ids = {
            course.class_id for course in Course.query.filter_by(institute_code=inst_code).all()
        }
        if not set(class_ids).issubset(valid_class_ids):
            flash("One or more selected classes are invalid.", "danger")
            return redirect(url_for("main.manage_subjects"))
        teacher_id = request.form.get("teacher_id", "").strip()
        teacher = Teacher.query.filter_by(institute_code=inst_code, teacher_id=teacher_id).first()
        if not teacher:
            flash("Selected teacher is invalid.", "danger")
            return redirect(url_for("main.manage_subjects"))
        subject_code = request.form.get("subject_code", "").strip()
        subject_name = request.form.get("subject_name", "").strip()
        if not subject_code or not subject_name:
            flash("Subject code and name are required.", "danger")
            return redirect(url_for("main.manage_subjects"))
        if Subject.query.filter_by(institute_code=inst_code, subject_code=subject_code).first():
            flash("That subject code already exists.", "danger")
            return redirect(url_for("main.manage_subjects"))
        try:
            session_len = _parse_form_int("session_length", 1, 8, 1)
            required_hours = _parse_form_int("required_hours", 1, 100)
            total_course_hours = _parse_form_int("total_course_hours", 1, 1000, 50)
        except ValueError as error:
            flash(str(error), "danger")
            return redirect(url_for("main.manage_subjects"))
        if required_hours % session_len:
            flash("Required weekly hours must be divisible by the session length.", "danger")
            return redirect(url_for("main.manage_subjects"))

        subject = Subject(
            institute_id=session.get("admin_id"),
            institute_code=inst_code,
            subject_code=subject_code,
            subject_name=subject_name,
            class_id=",".join(class_ids),
            teacher_id=teacher_id,
            teacher_id_fk=teacher.id,
            total_course_hours=total_course_hours,
            required_hours=required_hours,
            subject_type=request.form["subject_type"],
            session_length=session_len,
        )

        # Populate M2M Course mappings
        courses_db = Course.query.filter(Course.institute_code == inst_code, Course.class_id.in_(class_ids)).all()
        for course in courses_db:
            subject.courses.append(course)

        db.session.add(subject)
        db.session.commit()
        flash("Subject mapped!", "success")
        return redirect(url_for("main.manage_subjects"))

    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    courses = Course.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    weeks_setting = Settings.query.filter_by(
        institute_code=inst_code, key="weeks_per_semester"
    ).first()
    weeks_per_semester = int(weeks_setting.value) if weeks_setting else 15
    return render_template(
        "admin/manage_master.html",
        manage_type="subject",
        items=subjects,
        courses=courses,
        teachers=teachers,
        weeks_per_semester=weeks_per_semester,
    )


@main_bp.route("/edit_course/<int:id>", methods=["GET", "POST"])
@login_required_admin
def edit_course(id):
    course = _tenant_record_or_404(Course, id)
    if request.method == "POST":
        course.class_id = request.form["class_id"]
        course.department = request.form["department"]
        course.semester = request.form["semester"]
        course.division = request.form["division"]
        db.session.commit()
        flash("Course updated!", "success")
        return redirect(url_for("main.manage_courses"))
    return render_template("admin/edit_master.html", item=course, edit_type="course")


@main_bp.route("/edit_teacher/<int:id>", methods=["GET", "POST"])
@login_required_admin
def edit_teacher(id):
    teacher = _tenant_record_or_404(Teacher, id)
    if request.method == "POST":
        teacher.teacher_id = request.form["teacher_id"]
        teacher.name = request.form["name"]
        teacher.email = request.form["email"]
        teacher.departments = request.form["departments"]
        teacher.max_hours = request.form["max_hours"]
        days = request.form.getlist("days")
        if days:
            teacher.available_days = ",".join(days)
        db.session.commit()
        flash("Teacher updated!", "success")
        return redirect(url_for("main.manage_teachers"))

    courses = Course.query.filter_by(institute_code=session["institute_code"]).all()
    depts = list(set([c.department for c in courses]))
    return render_template(
        "admin/edit_master.html", item=teacher, edit_type="teacher", unique_depts=depts
    )


@main_bp.route("/edit_subject/<int:id>", methods=["GET", "POST"])
@login_required_admin
def edit_subject(id):
    subject = _tenant_record_or_404(Subject, id)
    if request.method == "POST":
        subject_code = request.form.get("subject_code", "").strip()
        duplicate = Subject.query.filter(
            Subject.institute_code == session["institute_code"],
            Subject.subject_code == subject_code,
            Subject.id != subject.id,
        ).first()
        if duplicate:
            flash("That subject code already exists.", "danger")
            return redirect(url_for("main.edit_subject", id=id))
        class_ids = request.form.getlist("class_id")
        valid_class_ids = {
            course.class_id
            for course in Course.query.filter_by(institute_code=session["institute_code"]).all()
        }
        teacher_id = request.form.get("teacher_id", "").strip()
        if not class_ids or not set(class_ids).issubset(valid_class_ids):
            flash("Select at least one valid class.", "danger")
            return redirect(url_for("main.edit_subject", id=id))
        teacher_db = Teacher.query.filter_by(
            institute_code=session["institute_code"], teacher_id=teacher_id
        ).first()
        if not teacher_db:
            flash("Selected teacher is invalid.", "danger")
            return redirect(url_for("main.edit_subject", id=id))
        try:
            total_course_hours = _parse_form_int("total_course_hours", 1, 1000, 50)
            required_hours = _parse_form_int("required_hours", 1, 100)
            session_len = _parse_form_int("session_length", 1, 8, 1)
        except ValueError as error:
            flash(str(error), "danger")
            return redirect(url_for("main.edit_subject", id=id))
        if required_hours % session_len:
            flash("Required weekly hours must be divisible by the session length.", "danger")
            return redirect(url_for("main.edit_subject", id=id))

        subject.subject_code = subject_code
        subject.subject_name = request.form.get("subject_name", "").strip()
        subject.class_id = ",".join(class_ids)
        subject.teacher_id = teacher_id
        subject.teacher_id_fk = teacher_db.id
        subject.total_course_hours = total_course_hours
        subject.required_hours = required_hours
        subject.subject_type = request.form.get("subject_type", "Theory")
        subject.session_length = session_len

        subject.courses = []
        courses_db = Course.query.filter(Course.institute_code == session["institute_code"], Course.class_id.in_(class_ids)).all()
        for course in courses_db:
            subject.courses.append(course)

        db.session.commit()
        flash("Subject updated!", "success")
        return redirect(url_for("main.manage_subjects"))

    courses = Course.query.filter_by(institute_code=session["institute_code"]).all()
    teachers = Teacher.query.filter_by(institute_code=session["institute_code"]).all()
    all_settings = Settings.query.filter_by(institute_code=session["institute_code"]).all()
    settings_dict = {s.key: s.value for s in all_settings}
    return render_template(
        "admin/edit_master.html",
        item=subject,
        edit_type="subject",
        courses=courses,
        teachers=teachers,
        settings=settings_dict,
    )


@main_bp.route("/delete/<resource_type>/<int:id>", methods=["POST"])
@login_required_admin
def delete_item(resource_type, id):
    resources = {
        "course": (Course, "manage_courses"),
        "teacher": (Teacher, "manage_teachers"),
        "subject": (Subject, "manage_subjects"),
    }
    if resource_type not in resources:
        return "Unsupported resource type", 404
    model, route = resources[resource_type]
    item = _tenant_record_or_404(model, id)
    db.session.delete(item)
    db.session.commit()
    flash("Deleted successfully.", "info")
    return redirect(url_for("main." + route))


@main_bp.route("/bulk_delete/<type>", methods=["POST"])
@login_required_admin
def bulk_delete_items(type):
    inst_code = session["institute_code"]
    if type not in {"course", "teacher", "subject"}:
        return "Unsupported resource type", 404
    selected_ids = request.form.getlist("selected_ids")

    if not selected_ids:
        flash("No items selected for deletion.", "warning")
        if type == "course":
            return redirect(url_for("main.manage_courses"))
        if type == "teacher":
            return redirect(url_for("main.manage_teachers"))
        return redirect(url_for("main.manage_subjects"))

    try:
        if type == "course":
            num_deleted = Course.query.filter(
                Course.id.in_(selected_ids), Course.institute_code == inst_code
            ).delete(synchronize_session=False)
            route = "manage_courses"
        elif type == "teacher":
            num_deleted = Teacher.query.filter(
                Teacher.id.in_(selected_ids), Teacher.institute_code == inst_code
            ).delete(synchronize_session=False)
            route = "manage_teachers"
        elif type == "subject":
            num_deleted = Subject.query.filter(
                Subject.id.in_(selected_ids), Subject.institute_code == inst_code
            ).delete(synchronize_session=False)
            route = "manage_subjects"

        db.session.commit()
        flash(f"Successfully deleted {num_deleted} items.", "info")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Bulk deletion failed")
        flash("Unable to delete the selected items. No changes were saved.", "danger")
        route = "admin_dash"

    return redirect(url_for("main." + route))


@main_bp.route("/generate_timetable", methods=["GET"])
@login_required_admin
def generate_timetable():
    admin = db.session.get(Institute, session["admin_id"])
    inst_code = session["institute_code"]

    # Render the interactive generation UI
    return render_template("admin/generate_timetable.html", admin=admin)


@main_bp.route("/api/generate_timetable", methods=["POST"])
@login_required_admin
def api_generate_timetable():
    inst_code = session["institute_code"]
    from utils.timetable_adapter import engine_generate_timetable

    try:
        result = engine_generate_timetable(inst_code)
    except Exception as e:
        import traceback
        from flask import current_app
        current_app.logger.error("Timetable generation failed with exception: %s\n%s", str(e), traceback.format_exc())
        result = {
            "success": False,
            "status": "FAILED",
            "message": "An internal server error occurred while generating the timetable. Please check server logs.",
            "stats": {},
            "diagnostics": {
                "status": "FAILED",
                "reason_code": "INTERNAL_ERROR",
                "primary_bottleneck": "Internal Error",
                "affected_courses": [],
                "affected_subjects": [],
                "affected_teachers": [],
                "required_capacity": 0,
                "available_capacity": 0,
                "shortage": 0,
                "suggestions": ["Please review the timetable configuration and try again."]
            }
        }
    return jsonify(result)


@main_bp.route("/generation_history")
@login_required_admin
def generation_history():
    admin = db.session.get(Institute, session["admin_id"])
    inst_code = session["institute_code"]
    history = (
        GenerationHistory.query.filter_by(institute_code=inst_code)
        .order_by(GenerationHistory.created_at.desc())
        .limit(50)
        .all()
    )
    return render_template("admin/generation_history.html", admin=admin, history=history)


@main_bp.route("/clear_history", methods=["POST"])
@login_required_admin
def clear_history():
    inst_code = session["institute_code"]
    GenerationHistory.query.filter_by(institute_code=inst_code).delete()
    db.session.commit()
    flash("Generation history cleared successfully.", "success")
    return redirect(url_for("main.generation_history"))


@main_bp.route("/view_timetable")
@login_required_admin
def view_timetable():
    inst_code = session["institute_code"]
    from utils.helpers import ScheduleConfig, get_local_date

    courses = Course.query.filter_by(institute_code=inst_code).all()
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    selected_class = request.args.get("class_id")
    date_str = request.args.get("date")

    if date_str:
        try:
            from datetime import datetime
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_date = get_local_date()
    else:
        selected_date = get_local_date()

    schedule = {}

    schedule_config = ScheduleConfig(inst_code)
    days = schedule_config.working_days
    time_slots = schedule_config.get_dynamic_time_slots()
    lunch_after = schedule_config.lunch_after
    break_duration = schedule_config.break_duration

    day_dates = {}

    if selected_class:
        if not Course.query.filter_by(institute_code=inst_code, class_id=selected_class).first():
            flash("Selected class was not found.", "warning")
            return redirect(url_for("main.view_timetable"))
        from utils.timetable_adapter import get_live_week_timetable

        filters = {"class_id": selected_class}
        live_week = get_live_week_timetable(inst_code, reference_date=selected_date, filters=filters)
        days = live_week["working_days"]
        day_dates = live_week["day_dates"]

        schedule = {day: {} for day in days}
        for entry in live_week["records"]:
            if entry.day_name in schedule:
                schedule[entry.day_name][entry.start_time] = entry
        time_slots = trim_time_slots(schedule, time_slots, lunch_after)

    inst = Institute.query.filter_by(institute_code=inst_code).first()
    inst_name = inst.name if inst else "Institute"

    # Fetch teachers and subjects for manual editing
    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    if selected_class:
        subjects = [
            subject
            for subject in Subject.query.filter_by(institute_code=inst_code).all()
            if selected_class in {value.strip() for value in subject.class_id.split(",")}
        ]
    else:
        subjects = []

    days_data = build_timetable_view_model(schedule, days, time_slots)
    return render_template(
        "shared/view_timetable.html",
        courses=courses,
        selected_class=selected_class,
        selected_date=date_str,
        schedule=schedule,
        days_data=days_data,
        days=days,
        day_dates=day_dates,
        time_slots=time_slots,
        lunch_after=lunch_after,
        break_duration=break_duration,
        inst_name=inst_name,
        teachers=teachers,
        subjects=subjects,
    )


@main_bp.route("/api/get_slot_data", methods=["GET"])
@login_required_admin
def get_slot_data():
    inst_code = session.get("institute_code")
    day = request.args.get("day")
    start_time = request.args.get("start_time")
    class_id = request.args.get("class_id")

    if not all([inst_code, day, start_time, class_id]):
        return jsonify({"error": "Missing parameters"}), 400

    from utils.helpers import ScheduleConfig
    if day not in ScheduleConfig(inst_code).working_days:
        return jsonify({"error": "Invalid day"}), 400
    if not Course.query.filter_by(institute_code=inst_code, class_id=class_id).first():
        return jsonify({"error": "Class not found"}), 404

    teachers = Teacher.query.filter_by(institute_code=inst_code).all()
    busy_entries = Timetable.query.filter_by(
        institute_code=inst_code,
        day_name=day,
        start_time=start_time,
    ).all()
    busy_by_teacher = {entry.teacher_name: entry for entry in busy_entries}

    subjects_by_teacher = {teacher.teacher_id: [] for teacher in teachers}
    subjects = Subject.query.filter_by(institute_code=inst_code).all()
    for subject in subjects:
        class_ids = {value.strip() for value in subject.class_id.split(",")}
        if class_id in class_ids and subject.teacher_id in subjects_by_teacher:
            subjects_by_teacher[subject.teacher_id].append(subject.subject_name)

    result = {"free": [], "busy": []}

    for t in teachers:
        busy_entry = busy_by_teacher.get(t.name)
        t_data = {
            "name": t.name,
            "subjects": subjects_by_teacher[t.teacher_id],
            "busy_class": busy_entry.class_id if busy_entry else None,
        }

        if busy_entry:
            result["busy"].append(t_data)
        else:
            result["free"].append(t_data)

    return jsonify(result)


@main_bp.route("/edit_timetable_slot", methods=["POST"])
@login_required_admin
def edit_timetable_slot():

    entry_id = request.form.get("entry_id")
    new_subject = request.form.get("new_subject")
    new_teacher = request.form.get("new_teacher")

    entry = Timetable.query.filter_by(
        id=entry_id,
        institute_code=session["institute_code"],
    ).first()
    if not entry:
        flash("Timetable slot not found.", "danger")
        return redirect(url_for("main.view_timetable", class_id=request.form.get("class_id")))

    teacher = Teacher.query.filter_by(
        institute_code=entry.institute_code,
        name=new_teacher,
    ).first()
    if not teacher:
        flash("Selected teacher does not belong to this institute.", "danger")
        return redirect(url_for("main.view_timetable", class_id=entry.class_id))

    subject = Subject.query.filter_by(
        institute_code=entry.institute_code,
        subject_name=new_subject,
    ).first()
    if not subject or entry.class_id not in {
        value.strip() for value in subject.class_id.split(",")
    }:
        flash("Selected subject is not mapped to this class.", "danger")
        return redirect(url_for("main.view_timetable", class_id=entry.class_id))

    if new_teacher != entry.teacher_name:
        clash = Timetable.query.filter_by(
            institute_code=entry.institute_code,
            day_name=entry.day_name,
            start_time=entry.start_time,
            teacher_name=new_teacher,
        ).first()

        if clash:
            flash(
                f"Cannot assign {new_teacher}. They are already teaching Class {clash.class_id} at this time!",
                "danger",
            )
            return redirect(url_for("main.view_timetable", class_id=entry.class_id))

    # Check available days hard constraint
    if teacher:
        from utils.helpers import ScheduleConfig
        t_days = (
            [d.strip() for d in (teacher.available_days or "").split(",")]
            if teacher.available_days
            else ScheduleConfig(entry.institute_code).working_days
        )
        if entry.day_name not in t_days:
            flash(
                f"Cannot assign {new_teacher}. They are not available on {entry.day_name}.",
                "danger",
            )
            return redirect(url_for("main.view_timetable", class_id=entry.class_id))

    # Update entry
    entry.subject_name = new_subject
    entry.teacher_name = new_teacher

    # If the entry was already a date-specific proxy override (leave_id is not null),
    # we preserve is_proxy=True and leave_id so it cleans up properly.
    # Otherwise, it's a permanent override.
    if getattr(entry, "specific_date", None) is None:
        if hasattr(entry, "is_proxy"):
            entry.is_proxy = False
            entry.original_teacher = None

    db.session.commit()
    flash("Timetable slot updated successfully!", "success")
    return redirect(url_for("main.view_timetable", class_id=entry.class_id))


@main_bp.route("/manual_assign_slot", methods=["POST"])
@login_required_admin
def manual_assign_slot():
    inst_code = session["institute_code"]
    class_id = request.form.get("class_id")
    day_name = request.form.get("day_name")
    start_time = request.form.get("start_time")
    assign_type = request.form.get("assign_type")  # 'lecture' or 'event'

    if not all([class_id, day_name, start_time, assign_type]):
        flash("Missing required fields for assignment.", "danger")
        return redirect(url_for("main.view_timetable", class_id=class_id))
    from utils.helpers import ScheduleConfig
    if day_name not in ScheduleConfig(inst_code).working_days:
        flash("Invalid timetable day.", "danger")
        return redirect(url_for("main.view_timetable", class_id=class_id))
    if not Course.query.filter_by(institute_code=inst_code, class_id=class_id).first():
        flash("Selected class does not belong to this institute.", "danger")
        return redirect(url_for("main.view_timetable"))
    if Timetable.query.filter_by(
        institute_code=inst_code,
        class_id=class_id,
        day_name=day_name,
        start_time=start_time,
        specific_date=None,
    ).first():
        flash("This class already has an assignment in the selected slot.", "danger")
        return redirect(url_for("main.view_timetable", class_id=class_id))

    end_time = None
    # Figure out end time from dynamic slots
    slots = get_dynamic_time_slots(inst_code)
    for s in slots:
        if s[0] == start_time:
            end_time = s[1]
            break

    if not end_time:
        flash("Invalid start time.", "danger")
        return redirect(url_for("main.view_timetable", class_id=class_id))

    subject_name = ""
    teacher_name = ""
    is_proxy = False

    if assign_type == "lecture":
        subject_name = request.form.get("new_subject")
        teacher_name = request.form.get("new_teacher")
        if not subject_name or not teacher_name:
            flash("Teacher and Subject are required for lecture assignment.", "danger")
            return redirect(url_for("main.view_timetable", class_id=class_id))

        subject = Subject.query.filter_by(
            institute_code=inst_code,
            subject_name=subject_name,
        ).first()
        if not subject or class_id not in {value.strip() for value in subject.class_id.split(",")}:
            flash("Selected subject is not mapped to this class.", "danger")
            return redirect(url_for("main.view_timetable", class_id=class_id))

        t_obj = Teacher.query.filter_by(
            institute_code=inst_code,
            name=teacher_name,
        ).first()
        if not t_obj:
            flash("Selected teacher does not belong to this institute.", "danger")
            return redirect(url_for("main.view_timetable", class_id=class_id))

        clash = Timetable.query.filter_by(
            institute_code=inst_code,
            day_name=day_name,
            start_time=start_time,
            teacher_name=teacher_name,
        ).first()

        if clash:
            flash(
                f"Cannot assign {teacher_name}. They are already assigned to Class {clash.class_id} at this time!",
                "danger",
            )
            return redirect(url_for("main.view_timetable", class_id=class_id))

        # Check available days hard constraint
        if t_obj:
            t_days = (
                [d.strip() for d in (t_obj.available_days or "").split(",")]
                if t_obj.available_days
                else ScheduleConfig(inst_code).working_days
            )
            if day_name not in t_days:
                flash(
                    f"Cannot assign {teacher_name}. They are not available on {day_name}.", "danger"
                )
                return redirect(url_for("main.view_timetable", class_id=class_id))

    elif assign_type == "event":
        subject_name = request.form.get("event_name")
        teacher_name = "Event/Workshop"
        if not subject_name:
            flash("Event name is required.", "danger")
            return redirect(url_for("main.view_timetable", class_id=class_id))

    # Add the entry
    new_entry = Timetable(
        institute_code=inst_code,
        class_id=class_id,
        day_name=day_name,
        start_time=start_time,
        end_time=end_time,
        subject_name=subject_name,
        teacher_name=teacher_name,
        is_proxy=is_proxy,
    )
    db.session.add(new_entry)
    db.session.commit()

    flash("Timetable slot assigned successfully!", "success")
    return redirect(url_for("main.view_timetable", class_id=class_id))


@main_bp.route("/export_timetables")
@login_required_admin
def export_timetables():
    inst_code = session["institute_code"]

    # Fetch Classes
    classes = (
        db.session.query(Timetable.class_id).filter_by(institute_code=inst_code).distinct().all()
    )
    class_ids = [c[0] for c in classes]

    if not class_ids:
        flash("No timetable generated yet to export!", "warning")
        return redirect(url_for("main.view_timetable"))

    # Fetch Settings & Slots
    from utils.helpers import ScheduleConfig
    schedule_config = ScheduleConfig(inst_code)
    settings = Settings.query.filter_by(institute_code=inst_code).all()
    s = {st.key: st.value for st in settings}

    institute_name = s.get("institute_name", "INSTITUTE TIMETABLE").upper()
    lunch_after = schedule_config.lunch_after

    # Time slots format: "08:00 AM - 09:00 AM"
    time_slots = schedule_config.get_dynamic_time_slots()
    slots = [f"{slot[0]} - {slot[1]}" for slot in time_slots]
    days = schedule_config.working_days

    # Excel Styling Variables
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    title_font = Font(name="Arial", size=16, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    cell_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    lunch_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cls in class_ids:
        safe_sheet_name = str(cls).replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(title=safe_sheet_name)

        # Titles
        c1 = ws.cell(row=1, column=1, value=institute_name)
        c1.font = title_font
        c1.alignment = center_align
        c2 = ws.cell(row=2, column=1, value=f"TIMETABLE / SCHEDULE - {cls}")
        c2.font = header_font
        c2.alignment = center_align
        c2.fill = blue_fill

        ws.cell(row=3, column=1, value="Day").font = header_font
        ws.cell(row=3, column=1).alignment = center_align
        ws.cell(row=3, column=1).border = thin_border
        ws.column_dimensions["A"].width = 12

        col_idx = 2
        lunch_printed = False
        slot_to_col = {}

        # Build Header & Lunch Break Column
        for i, slot in enumerate(slots):
            # Print Lunch Column exactly after the specified lecture
            if not lunch_printed and i == lunch_after:
                lc = ws.cell(row=3, column=col_idx, value="LUNCH\nBREAK")
                lc.font = header_font
                lc.alignment = center_align
                lc.border = thin_border
                lc.fill = lunch_fill
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

                # Merge and rotate text vertically
                ws.merge_cells(
                    start_row=4, start_column=col_idx, end_row=4 + len(days) - 1, end_column=col_idx
                )
                lb_cell = ws.cell(row=4, column=col_idx, value="L U N C H   B R E A K")
                lb_cell.alignment = Alignment(
                    textRotation=90, horizontal="center", vertical="center", wrap_text=True
                )
                lb_cell.font = Font(name="Arial", size=14, bold=True, color="595959")
                lb_cell.fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                )

                for r in range(4, 4 + len(days)):
                    ws.cell(row=r, column=col_idx).border = thin_border
                col_idx += 1
                lunch_printed = True

            # Print Time Slot Header
            slot_to_col[slot] = col_idx
            cell = ws.cell(row=3, column=col_idx, value=slot.replace(" - ", "\n"))
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
            col_idx += 1

        max_col = col_idx - 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

        # Fetch data for this specific class
        tt_entries = Timetable.query.filter_by(institute_code=inst_code, class_id=cls).all()
        tt_dict = {
            (e.day_name, f"{e.start_time} - {e.end_time}"): f"{e.subject_name}\n({e.teacher_name})"
            for e in tt_entries
        }

        # Populate Grid Data
        for row_idx, day in enumerate(days, start=4):
            day_cell = ws.cell(row=row_idx, column=1, value=day)
            day_cell.font = header_font
            day_cell.alignment = center_align
            day_cell.border = thin_border

            for slot in slots:
                c_idx = slot_to_col[slot]
                cell = ws.cell(row=row_idx, column=c_idx, value=tt_dict.get((day, slot), "---"))
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border

    # Save Excel to memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="AutoTime_Master_Schedule.xlsx", as_attachment=True)


@main_bp.route("/college_settings", methods=["GET", "POST"])
@login_required_admin
def college_settings():
    inst_code = session["institute_code"]

    if request.method == "POST":
        try:
            values = {
                "total_lectures": str(_parse_form_int("total_lectures", 1, 20)),
                "lecture_duration": str(_parse_form_int("lecture_duration", 15, 240)),
                "break_time": str(_parse_form_int("break_time", 0, 180)),
                "lunch_after_lecture": str(_parse_form_int("lunch_after_lecture", 1, 20)),
                "weeks_per_semester": str(_parse_form_int("weeks_per_semester", 1, 52)),
            }
            start_time = request.form.get("start_time", "")
            datetime.strptime(start_time, "%H:%M")
            values["start_time"] = start_time

            working_days_list = request.form.getlist("working_days")
            if not working_days_list:
                working_days_list = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
            values["working_days"] = ",".join(working_days_list)
        except ValueError as error:
            flash(str(error) or "Invalid institute settings.", "danger")
            return redirect(url_for("main.college_settings"))
        if int(values["lunch_after_lecture"]) > int(values["total_lectures"]):
            flash("Lunch break cannot start after the final lecture.", "danger")
            return redirect(url_for("main.college_settings"))

        for key, value in values.items():
            setting = Settings.query.filter_by(institute_code=inst_code, key=key).first()
            if setting:
                setting.value = value
            else:
                db.session.add(Settings(institute_code=inst_code, key=key, value=value))

        db.session.commit()
        flash("College Settings Updated Successfully!", "success")
        return redirect(url_for("main.college_settings"))

    all_settings = Settings.query.filter_by(institute_code=inst_code).all()
    settings_dict = {s.key: s.value for s in all_settings}

    return render_template("admin/college_settings.html", settings=settings_dict)


@main_bp.route("/admin/requests/approve/<int:req_id>", methods=["POST"])
@login_required_admin
def approve_teacher_request(req_id):
    req = TeacherUpdateRequest.query.get_or_404(req_id)
    if req.institute_code != session.get("institute_code"):
        return "Unauthorized", 403

    teacher = Teacher.query.filter_by(
        teacher_id=req.teacher_id, institute_code=req.institute_code
    ).first()
    if teacher:
        if req.new_name:
            old_name = teacher.name
            teacher.name = req.new_name
            # Cascade name changes to denormalized fields
            Timetable.query.filter_by(institute_code=req.institute_code, teacher_name=old_name).update({"teacher_name": req.new_name})
            TeacherLeave.query.filter_by(institute_code=req.institute_code, teacher_name=old_name).update({"teacher_name": req.new_name})
            # If there's any other model using teacher_name as string, update it here.

        if req.new_email:
            email_owner = Teacher.query.filter(
                Teacher.email == req.new_email,
                Teacher.id != teacher.id,
            ).first()
            if (
                email_owner
                or Institute.query.filter_by(admin_email=req.new_email).first()
                or Student.query.filter_by(email=req.new_email).first()
            ):
                flash("That email address is already in use.", "danger")
                return redirect(url_for("main.admin_dash"))
            teacher.email = req.new_email
        req.status = "Approved"
        db.session.commit()
        flash("Teacher update request approved.", "success")
    return redirect(url_for("main.admin_dash"))


@main_bp.route("/admin/requests/reject/<int:req_id>", methods=["POST"])
@login_required_admin
def reject_teacher_request(req_id):
    req = TeacherUpdateRequest.query.get_or_404(req_id)
    if req.institute_code != session.get("institute_code"):
        return "Unauthorized", 403

    req.status = "Rejected"
    db.session.commit()
    flash("Teacher update request rejected.", "info")
    return redirect(url_for("main.admin_dash"))


from models import AcademicCalendar, Notification
from datetime import datetime


@main_bp.route("/manage_calendar", methods=["GET", "POST"])
@login_required_admin
def manage_calendar():
    inst_code = session["institute_code"]

    if request.method == "POST":
        date_str = request.form.get("date")
        event_name = request.form.get("event_name")
        department = request.form.get("department", "All")  # 'All' or specific dept
        is_holiday = request.form.get("is_holiday") == "on"

        try:
            event_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            new_event = AcademicCalendar(
                institute_code=inst_code,
                date=event_date,
                event_name=event_name,
                department=department,
                is_holiday=is_holiday,
            )
            db.session.add(new_event)
            db.session.commit()

            # Holidays don't need proxies, they cancel classes.

            flash("Event added successfully! Proxy Engine ran for affected lectures.", "success")
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Academic calendar event creation failed")
            flash("Unable to add the event. No changes were saved.", "danger")

        return redirect(url_for("main.manage_calendar"))

    events = (
        AcademicCalendar.query.filter_by(institute_code=inst_code)
        .order_by(AcademicCalendar.date)
        .all()
    )
    # Fetch unique departments from Courses
    depts = [
        r.department
        for r in db.session.query(Course.department).filter_by(institute_code=inst_code).distinct()
    ]
    return render_template("admin/manage_calendar.html", events=events, depts=depts)


@main_bp.route("/notifications")
@login_required_admin
def admin_notifications():
    inst_code = session["institute_code"]
    notifs = (
        Notification.query.filter_by(institute_code=inst_code, user_type="admin")
        .order_by(Notification.created_at.desc())
        .all()
    )

    # Mark as read
    unread = [n for n in notifs if not n.is_read]
    if unread:
        for n in unread:
            n.is_read = True
        db.session.commit()

    return render_template("admin/notifications.html", notifications=notifs)


@main_bp.route("/clear_admin_notifications", methods=["POST"])
@login_required_admin
def clear_admin_notifications():
    inst_code = session["institute_code"]
    Notification.query.filter_by(institute_code=inst_code, user_type="admin").delete()
    db.session.commit()
    flash("All notifications cleared.", "success")
    return redirect(url_for("main.admin_notifications"))


@main_bp.route("/admin/leave_requests")
@login_required_admin
def leave_requests():
    inst_code = session["institute_code"]
    pending = (
        TeacherLeave.query.filter_by(institute_code=inst_code, status="Pending")
        .order_by(TeacherLeave.date)
        .all()
    )
    history = (
        TeacherLeave.query.filter(
            TeacherLeave.institute_code == inst_code, TeacherLeave.status != "Pending"
        )
        .order_by(TeacherLeave.date.desc())
        .limit(30)
        .all()
    )

    pending_leaves = []
    for l in pending:
        t = Teacher.query.filter_by(institute_code=inst_code, teacher_id=l.teacher_id).first()
        if t:
            pending_leaves.append((l, t))

    history_leaves = []
    for l in history:
        t = Teacher.query.filter_by(institute_code=inst_code, teacher_id=l.teacher_id).first()
        if t:
            history_leaves.append((l, t))

    return render_template(
        "admin/leave_requests.html", pending_leaves=pending_leaves, history_leaves=history_leaves
    )


@main_bp.route("/admin/approve_leave/<int:leave_id>", methods=["POST"])
@login_required_admin
def approve_leave(leave_id):
    from utils.leave_service import approve_leave as service_approve_leave

    success, msg = service_approve_leave(
        leave_id,
        institute_code=session["institute_code"],
    )
    if success:
        flash(msg, "success")
    else:
        flash(msg, "danger")
    return redirect(url_for("main.leave_requests"))


@main_bp.route("/admin/reject_leave/<int:leave_id>", methods=["POST"])
@login_required_admin
def reject_leave(leave_id):
    leave = _tenant_record_or_404(TeacherLeave, leave_id)
    if leave.status != "Pending":
        flash("Only pending leaves can be rejected.", "warning")
        return redirect(url_for("main.leave_requests"))
    leave.status = "Rejected"
    db.session.commit()
    flash("Leave request rejected.", "success")
    return redirect(url_for("main.leave_requests"))


@main_bp.route("/admin/revoke_leave/<int:leave_id>", methods=["POST"])
@login_required_admin
def revoke_leave(leave_id):
    from utils.leave_service import cancel_leave

    success, msg = cancel_leave(
        leave_id,
        actor_name="Admin",
        institute_code=session["institute_code"],
    )
    if success:
        flash(msg, "success")
    else:
        flash(msg, "danger")
    return redirect(url_for("main.leave_requests"))
